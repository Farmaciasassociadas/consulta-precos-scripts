from collections import defaultdict
from pathlib import Path
from statistics import mean

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


CATALOGO = Path("PERFUMARIA CLASSIFICADO - COM CATEGORIAS.xlsx")
ENTRADAS = Path("Relatório notas fiscais 24-07_PERFUMARIA.xlsx")
SAIDA = Path("RECOMENDACAO_PRECOS_PERFUMARIA.xlsx")

# lucro líquido alvo e fator físico da política vigente; concorrência não foi fornecida.
POLITICA = {
    "PERFUMARIA": (18, 1.05),
    "PERFUMARIA > BRONZ & HIDRATANTES": (20, 1.06),
    "PERFUMARIA > COLONIAS & DESODORANTES": (18, 1.05),
    "PERFUMARIA > LINHA CAPILAR": (18, 1.05),
    "PERFUMARIA > DERMOCOSMETICOS": (15, 1.03),
    "PERFUMARIA > FRALDAS": (5, 1.00),
    "PERFUMARIA > HIGIENE BUCAL": (12, 1.02),
    "PERFUMARIA > HIGIENE PESSOAL": (12, 1.03),
    "PERFUMARIA > LINHA INFANTIL": (10, 1.01),
    "PERFUMARIA > MAQUIAGEM": (22, 1.08),
    "PERFUMARIA > TINTURAS": (15, 1.04),
}
GRADE = (0.49, 0.79, 0.90, 0.95, 0.99)


def ean(value):
    return str(value).strip().replace(".0", "") if value is not None else ""


def preco_grade(alvo, piso):
    candidatos = [i + fim for i in range(max(0, int(piso) - 1), int(alvo) + 3) for fim in GRADE]
    return round(min((p for p in candidatos if p >= piso - 1e-9), key=lambda p: (abs(p - alvo), p)), 2)


def custos_por_ean():
    ws = load_workbook(ENTRADAS, read_only=True, data_only=True).active
    saida = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        codigo, quantidade, total, filho, pai = ean(row[9]), row[13], row[25], str(row[26] or ""), str(row[27] or "")
        if "MARCA PROPRIA" in (filho.upper(), pai.upper()) or not codigo or not quantidade or total is None:
            continue
        custo = float(total) / float(quantidade)
        saida[codigo].append((custo, float(quantidade), float(total)))
    return saida


def eans_marca_propria():
    ws = load_workbook(ENTRADAS, read_only=True, data_only=True).active
    return {ean(row[9]) for row in ws.iter_rows(min_row=2, values_only=True)
            if "MARCA PROPRIA" in (str(row[26] or "").upper(), str(row[27] or "").upper())}


def preparar(ws):
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for i, col in enumerate(ws.columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(45, max(12, max(len(str(c.value or "")) for c in col) + 2))


def main():
    custos = custos_por_ean()
    marca_propria = eans_marca_propria()
    catalogo = load_workbook(CATALOGO, read_only=True, data_only=True)["Planilha1"]
    resultado, manual, conferir = [], [], []
    for linha in catalogo.iter_rows(min_row=2, values_only=True):
        codigo, descricao, categoria = ean(linha[0]), linha[1], linha[12]
        compras = custos.get(codigo, [])
        if codigo in marca_propria:
            registro = [codigo, descricao, categoria, None, "N/A: marca própria excluída", 0, None, None, 0, None, None, None, None, None, None, None, None, "BAIXA", "REVISAO_MANUAL_MARCA_PROPRIA", "Marca própria: preço definido manualmente"]
            manual.append(registro)
            continue
        if not compras:
            registro = [codigo, descricao, categoria, None, "SEM_CUSTO_VALIDADO", 0, None, None, 0, None, None, None, None, None, None, None, None, "BAIXA", "REVISAO_MANUAL_SEM_CUSTO_VALIDADO", "EAN sem custo na fonte de entradas"]
            manual.append(registro)
            conferir.append(registro[:6])
            continue
        custo = mean(x[0] for x in compras)
        if categoria not in POLITICA:
            registro = [codigo, descricao, categoria, custo, "OK: valor total ÷ quantidade", len(compras), min(x[0] for x in compras), max(x[0] for x in compras), 0, None, None, None, None, None, None, None, None, "BAIXA", "REVISAO_MANUAL_SEM_MARKUP", "Categoria sem política"]
            manual.append(registro)
            continue
        lucro, fator = POLITICA[categoria]
        piso = custo / 0.7402
        alvo = custo / (1 - .025 - .0598 - .175 - lucro / 100)
        sugerido = preco_grade(alvo, piso)
        registro = [codigo, descricao, categoria, custo, "OK: valor total ÷ quantidade", len(compras), min(x[0] for x in compras), max(x[0] for x in compras), 0, None, None, "PROTECAO_MARGEM", piso, alvo, sugerido, (sugerido / custo - 1), (sugerido - custo) / sugerido, "MEDIA", "RECOMENDADO_SEM_CONCORRENCIA", "Sem preço concorrente numérico; alvo econômico da categoria"]
        resultado.append(registro)

    wb = Workbook()
    resumo = wb.active
    resumo.title = "Resumo"
    resumo.append(["Indicador", "Valor"])
    resumo.append(["Itens recomendados", len(resultado)])
    resumo.append(["Itens em revisão manual", len(manual)])
    resumo.append(["Sem preços concorrentes numéricos", len(resultado)])
    resumo.append(["Custo total dos itens recomendados", sum(r[3] for r in resultado)])
    resumo.append(["Preço sugerido total (unitário)", sum(r[14] for r in resultado)])
    preparar(resumo)

    headers = ["EAN", "Descrição", "Classificação", "Custo unitário", "Validação do custo", "Nº compras", "Custo mínimo", "Custo máximo", "Concorrentes válidos", "Mediana web", "Mediana física estimada", "Tier", "Piso", "Alvo econômico", "Preço sugerido", "Markup", "Margem bruta", "Confiança", "Status", "Justificativa"]
    completa = wb.create_sheet("Precificação Completa")
    completa.append(headers)
    for row in resultado:
        completa.append(row)
    preparar(completa)
    for col in (4, 7, 8, 10, 11, 13, 14, 15):
        for cell in list(completa.columns)[col - 1][1:]: cell.number_format = 'R$ #,##0.00'
    for col in (16, 17):
        for cell in list(completa.columns)[col - 1][1:]: cell.number_format = '0.0%'

    revisao = wb.create_sheet("Revisão Manual")
    revisao.append(headers)
    for row in manual:
        revisao.append(row)
    preparar(revisao)

    notas = wb.create_sheet("Notas e Premissas")
    for row in [
        ["Premissa", "Valor"],
        ["Fonte de custo", ENTRADAS.name],
        ["Cálculo do custo", "Valor Total do Item ÷ Qtde. Unitária; média por EAN"],
        ["Cartão", "2,50%"], ["Simples (estimativa)", "5,98%"], ["Despesas fixas", "17,50%"],
        ["Piso absoluto", "Custo ÷ 0,7402 (markup 35,1%)"],
        ["Grade", ",49; ,79; ,90; ,95; ,99"],
        ["Concorrência", "A aba Planilha2 não contém valores numéricos; nenhuma mediana foi calculada."],
        ["Decisão pendente", "Revisar preço sugerido após importar preços concorrentes por EAN."],
        ["Marca própria", "Excluída do custo; qualquer EAN sem custo foi encaminhado à revisão manual."],
    ]: notas.append(row)
    preparar(notas)

    conferir_ws = wb.create_sheet("Custos a Conferir")
    conferir_ws.append(["EAN", "Descrição", "Classificação", "Custo unitário", "Validação", "Nº compras"])
    for row in conferir: conferir_ws.append(row)
    preparar(conferir_ws)
    wb.save(SAIDA)
    print(f"Criado: {SAIDA.resolve()}")
    print(f"Recomendados: {len(resultado)}; revisão manual: {len(manual)}")


if __name__ == "__main__":
    assert preco_grade(15.2, 14.0) >= 14.0
    main()
