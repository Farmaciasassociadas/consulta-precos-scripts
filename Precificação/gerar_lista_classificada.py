from collections import defaultdict
from math import isfinite
from pathlib import Path
import re
import unicodedata

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path.cwd()
ITEMS = next(ROOT.glob("*produtos_com_precos_unitarios.xlsx"))
REPORT = next(ROOT.glob("*com_custo_unitario.xlsx"))
PRIOR = next(ROOT.glob("RECOMENDACAO_PRECOS_PERFUMARIA.xlsx"))
OUTPUT = ROOT / "LISTA_CLASSIFICADA_COM_PRECOS_UNITARIOS.xlsx"
TAXONOMY = Path(r"C:\Users\docze\.codex\skills\drogaria-precificacao\references\GRUPOS E SUBGRUPOS.xlsx")


def ean(value):
    return "".join(re.findall(r"\d", str(value or "")))


def text(value):
    value = unicodedata.normalize("NFD", str(value or "")).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", value.lower()).strip()


def is_vitamin(description):
    return bool(re.search(r"\bvit(?:amina)?\b|\bomega\b|\bcomplexo b\b|\bmultivit", text(description)))


def exclusive(description):
    return "EXCLUSIVOS > VITAMINAS" if is_vitamin(description) else "EXCLUSIVOS > GERAL"


def perfumaria(description):
    value = text(description)
    if re.search(r"\bfralda|\bpants\b|plenitud|roupa intima", value): return "PERFUMARIA > FRALDAS"
    if re.search(r"tintura|colorac|tonaliz|henna|agua oxig|oxidante|revelador|cor e ton", value): return "PERFUMARIA > TINTURAS"
    if re.search(r"maquiagem|\bbatom\b|\bgloss\b|esmalte|\bbase\b|corretivo|\bblush\b|rimel|cilios|delineador|\bsombra\b|lip oil|sobrancelha", value): return "PERFUMARIA > MAQUIAGEM"
    if re.search(r"creme dental|pasta dental|\bdental\b|escova (?:de )?dente|fio dental|enxaguante bucal|bochecho|higiene bucal|oral[- ]?b|colgate|listerine", value): return "PERFUMARIA > HIGIENE BUCAL"
    # Ponytail: 'baby' alone is not enough; Giovanna Baby is a fragrance line.
    if re.search(r"\bbebe\b|\bbaby\b|infantil|crianca|\bkids\b|johnson.?s baby|joao e maria|hipoglos", value) and ("giovanna baby" not in value or "kids" in value): return "PERFUMARIA > LINHA INFANTIL"
    if re.search(r"shampoo|\bsham\b|condicion|capilar|cabel[oa]|mascara capilar|creme (?:de )?pentear|leave[ -]?in|hair|pantene|siage", value): return "PERFUMARIA > LINHA CAPILAR"
    if re.search(r"desodorante|\bdes aero\b|\bds aero\b|\baer\.?\b|colonia|\bperfume\b|fragrancia|body splash|antitranspirante|roll[- ]?on|aerosol", value): return "PERFUMARIA > COLONIAS & DESODORANTES"
    if re.search(r"\bderma\b|acne|facial|serum|anti ?idade|retinol|vitamina c|hialuron|niacinamida|skincare|agua micelar|esfoliante facial|cerave|la roche|vichy|neutrogena|principia|eucerin", value): return "PERFUMARIA > DERMOCOSMETICOS"
    if re.search(r"protetor solar|\bfps\b|filtro solar|bronzeador|pos sol|hidrat|\bloc hid\b|locao corporal|oleo corporal|creme (?:para )?maos", value): return "PERFUMARIA > BRONZ & HIDRATANTES"
    return "PERFUMARIA > HIGIENE PESSOAL"


def map_source(parent, child, description):
    parent, child = text(parent), text(child)
    if parent == "perfumaria": return perfumaria(description)
    if parent == "generico":
        if "anticoncepcional" in child: return "GENERICO > ANTICONCEPCIONAL"
        if "antimicrobiano" in child: return "GENERICO > ANTIMICROBIANO"
        if "controlado" in child: return "GENERICO > CONTROLADO"
        return "GENERICO > USO CONTINUO" if "popular" in child else "GENERICO > O.T.C/MIP"
    if parent == "similar":
        if "anticoncepcional" in child: return "SIMILAR > ANTICONCEPCIONAL"
        return "SIMILAR > USO CONTINUO" if "popular" in child else "SIMILAR > O.T.C/MIP-SIMILAR"
    if parent == "referencia":
        if "anticoncepcional" in child: return "ETICOS > ANTICONCEPCIONAL"
        if "antimicrobiano" in child: return "ETICOS > ANTIMICROBIANO"
        if "controlado" in child: return "ETICOS > CONTROLADO"
        return "ETICOS > USO CONTINUO" if "popular" in child else "ETICOS > O.T.C/MIP"
    if parent == "liberado":
        if "bonificado" in child or "exclusiv" in text(description) or "santo habito" in text(description): return exclusive(description)
        if "eletron" in child or "acessor" in child: return "VAREJO > ACESSORIOS"
        if "nutricao" in child or "leites" in child: return "VAREJO > LEITES NUTRICAO"
        if "naturais" in child: return "VAREJO > FITOTERAPICOS"
        return "VAREJO > VAREJINHO"
    return "VAREJO > VAREJINHO"


def infer(description):
    value = text(description)
    if "santo habito" in value or "exclusiv" in value: return exclusive(description)
    if re.search(r"/gn\b|\bgen(?:erico)?\b", value):
        return "GENERICO > ANTICONCEPCIONAL" if re.search(r"anticoncepc|etinilestradiol|desogestrel|drospirenona", value) else "GENERICO > O.T.C/MIP"
    if re.search(r"leite|nan\b|nestogeno|ninho\b|nutren", value): return "VAREJO > LEITES NUTRICAO"
    if re.search(r"oximetro|nebulizador|termometro|tiras\b|lanceta|aparelho pressao|bomba tira", value): return "VAREJO > ACESSORIOS"
    if re.search(r"chocolate|bombom|pastilha|kit kat|bebida", value): return "VAREJO > BEBIDAS E BOMBONIERE"
    if re.search(r"vitamina|omega|colageno|suplement|whey|creatina", value): return "VAREJO > VITAMINAS"
    if re.search(r"mg\b|mcg\b|cpr\b|comp\b|caps?\b|xpe\b|gotas?\b|inj\b|ampola", value): return "ETICOS > O.T.C/MIP"
    return perfumaria(description)


def main():
    report_map = {}
    costs_by_ean = defaultdict(list)
    ws = load_workbook(REPORT, read_only=True, data_only=True).active
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = ean(row[9])
        if not code or not row[10]:
            continue
        if code not in report_map:
            report_map[code] = (row[28], row[27])
        try:
            quantity, stored_cost, total = float(row[13]), float(row[25]), float(row[26])
        except (TypeError, ValueError):
            continue
        if not all(isfinite(value) and value > 0 for value in (quantity, stored_cost, total)):
            continue
        # The report's total already includes freight and other item charges.
        # Do not reuse the corrupt values previously present in the product list.
        if abs(stored_cost * quantity - total) <= max(0.02, total * 0.000001):
            costs_by_ean[code].append(total / quantity)
    unit_cost = {code: sum(values) / len(values) for code, values in costs_by_ean.items()}

    allowed = {row[0] for row in load_workbook(TAXONOMY, read_only=True, data_only=True).active.iter_rows(values_only=True) if row[0]}

    # Earlier review output identifies products formerly removed as marca propria.
    exclusive_eans = set()
    for ws in load_workbook(PRIOR, read_only=True, data_only=True).worksheets:
        if ws.title == "Revisão Manual":
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[18] == "REVISAO_MANUAL_MARCA_PROPRIA": exclusive_eans.add(ean(row[0]))

    source = load_workbook(ITEMS, read_only=True, data_only=True).active
    out = Workbook(); sheet = out.active; sheet.title = "Produtos"
    sheet.append(["EAN", "Descrição do produto", "Preço unitário", "Classificação"])
    counts = defaultdict(int)
    seen = set()
    for row in source.iter_rows(values_only=True):
        code, description = ean(row[0]), row[1]
        if not code or code in seen: continue
        seen.add(code)
        if code in exclusive_eans:
            classification = exclusive(description)
        elif code in report_map:
            classification = map_source(*report_map[code], description)
        else:
            classification = infer(description)
        assert classification in allowed, classification
        price = unit_cost.get(code)
        sheet.append([code, description, price, classification]); counts[classification] += 1

    sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78")
    for cell in sheet["A"]: cell.number_format = "@"
    for cell in sheet["C"][1:]: cell.number_format = 'R$ #,##0.00'
    for column, width in {"A": 16, "B": 58, "C": 16, "D": 34}.items(): sheet.column_dimensions[column].width = width
    out.save(OUTPUT)

    assert sheet.max_row - 1 == 2437
    assert sum(counts.values()) == 2437
    assert all(price is None or 0 < price < 1000 for price in unit_cost.values())
    assert all(sheet.cell(row, 1).value for row in range(2, sheet.max_row + 1))
    print(f"{OUTPUT.resolve()}\nProdutos: {sheet.max_row - 1}\nExclusivos: {sum(value for key, value in counts.items() if key.startswith('EXCLUSIVOS'))}")


if __name__ == "__main__": main()
