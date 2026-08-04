"""Extrai os precos sugeridos da rodada oficial (v2) para os itens do estoque da DROGARIA.

Garante que a planilha da DROGARIA usa exatamente os valores da tabela matriz.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill

ESTOQUE_XLSX = Path(
    r"G:\.shortcut-targets-by-id\1q0IRmUp06SR55V7qNb7wVLwWjEauQntR\DROGARIA\estoque.xlsx"
)
SAIDA_XLSX = Path(__file__).parent.parent / "ESTOQUE_DROGARIA_PRECIFICADO_FROM_RODADA.xlsx"

SITES = [
    "drogaraia", "nissei", "saopaulo", "saojoao", "panvel",
    "sistema", "paguemenos", "precopopular", "drogariasp", "farmasp",
]


def ler_estoque_eans(caminho: Path) -> list[str]:
    """Le apenas os EANs validos do estoque.xlsx."""
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    eans = []
    for r in ws.iter_rows(values_only=True):
        if len(r) < 9:
            continue
        ean = r[8]
        if ean and isinstance(ean, str) and ean.strip().isdigit():
            eans.append(ean.strip())
    wb.close()
    return eans


def main() -> None:
    print("Lendo EANs do estoque.xlsx...")
    eans_estoque = set(ler_estoque_eans(ESTOQUE_XLSX))
    print(f"  {len(eans_estoque)} EANs unicos")

    conn = sqlite3.connect("precificador.db")
    conn.row_factory = sqlite3.Row
    rodada_id = conn.execute("SELECT MAX(id) FROM rodada").fetchone()[0]
    print(f"Lendo dados da rodada {rodada_id} (mais recente)...")

    recomendacoes = conn.execute(
        "SELECT * FROM recomendacao WHERE rodada_id = ?", (rodada_id,)
    ).fetchall()
    print(f"  {len(recomendacoes)} recomendacoes na rodada {rodada_id}")

    recomendacoes_drogaria = [
        r for r in recomendacoes
        if r["ean"] in eans_estoque or r["status"] == "MARCA_EXCLUSIVA_MANUAL"
    ]
    print(f"  {len(recomendacoes_drogaria)} pertencem ao estoque da DROGARIA (inclui marca exclusiva fora do estoque)")

    precos_por_ean = {}
    for ean in eans_estoque:
        linhas = conn.execute(
            "SELECT site, preco FROM preco_concorrente WHERE ean = ? AND status = 'OK' ORDER BY data_hora DESC",
            (ean,),
        ).fetchall()
        precos_por_ean[ean] = {site: preco for site, preco in linhas}

    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque Precificado (Rodada 9)"

    cabecalho = [
        "EAN", "Descricao", "Custo unitario", "Preco venda atual", "Preco sugerido", "Preco Brick (VUM)",
        *[s.upper() for s in SITES],
        "Categoria", "Natureza fiscal", "Tier", "Status", "Justificativa",
    ]
    ws.append(cabecalho)

    moeda_idx = {3, 4, 5, 6} | {7 + i for i in range(len(SITES))}
    for rec in recomendacoes_drogaria:
        precos_sites = precos_por_ean.get(rec["ean"], {})
        row = [
            rec["ean"],
            rec["descricao"],
            rec["custo"],
            rec["preco_atual"],
            rec["preco_sugerido"],
            rec["vum_brick"],
            *[precos_sites.get(s) for s in SITES],
            rec["categoria_provisoria"],
            rec["natureza_fiscal"],
            rec["tier"],
            rec["status"],
            rec["justificativa"],
        ]
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for idx in moeda_idx:
        for cell in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            cell[0].number_format = "R$ #,##0.00"
    for coluna in ws.columns:
        letra = coluna[0].column_letter
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=8)
        ws.column_dimensions[letra].width = min(largura + 2, 60)

    wb.save(SAIDA_XLSX)
    print(f"Exportado: {SAIDA_XLSX}")


if __name__ == "__main__":
    main()
