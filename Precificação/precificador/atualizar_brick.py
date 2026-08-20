r"""Substitui os dados do Brick (VUM, curva, segmento) pela planilha BRICK mais
recente do painel farmaceutico.

A planilha vem em quatro abas por eixo (RX, GEN, SIM, NMED) e o cabecalho nao
esta na mesma linha em todas -- por isso as colunas sao achadas pelo NOME, nunca
por indice. Atualiza `precificacao/dados/referencia_categoria_brick.csv`, que e'
a fonte do "Preco Brick" no MiniPreco e a ancora de mercado do motor.

    python atualizar_brick.py "C:\...\BRICK 1855 - PRODUTOS - MAT07_2026.xlsx"
    python atualizar_brick.py <xlsx> --dry-run
"""
from __future__ import annotations

import argparse
import csv
import shutil
import unicodedata
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

REFERENCIA = Path(r"C:\Users\docze\ConsultaPrecosEAN\precificacao\dados"
                  r"\referencia_categoria_brick.csv")
# `segmento` NAO e' o "Catalogo Guia" da planilha: no resto do projeto ele e' o
# EIXO (RX/GEN/SIM/NMED), que aqui vem do NOME DA ABA. Trocar isso muda o
# lucro-alvo do motor e o agrupamento do chamariz sem erro nenhum aparecer.
COLUNAS = {"ean": "ean", "vum": "vum", "curva em valor": "curva",
           "rk brick": "posicao"}
# Planilha corrente do painel farmaceutico, arquivada junto do consolidado.
BRICK_ATUAL = (Path(__file__).parent.parent / "outputs" / "consolidado_estoque"
               / "BRICK 1855 - PRODUTOS - MAT07_2026.xlsx")


def _chave(texto) -> str:
    bruto = unicodedata.normalize("NFKD", str(texto or ""))
    limpo = "".join(c for c in bruto if not unicodedata.combining(c))
    # A planilha vem com acento quebrado ("Cat?logo"): o "?" tambem cai fora.
    return " ".join("".join(c for c in limpo.lower() if c.isalnum() or c.isspace()).split())


def normalizar_ean(valor) -> str:
    """Mesma chave do resto do projeto: so' digitos, sem zero de enchimento."""
    digitos = "".join(c for c in str(valor or "") if c.isdigit())
    return digitos.lstrip("0") or digitos


def _numero(valor):
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str) and valor.strip():
        try:
            return float(valor.strip().replace(",", "."))
        except ValueError:
            return None
    return None


def ler_brick(caminho: Path) -> dict[str, dict]:
    """{ean: {vum, curva, segmento}} de todas as abas com cabecalho de produto."""
    wb = load_workbook(caminho, read_only=True, data_only=True)
    dados: dict[str, dict] = {}
    for aba in wb.worksheets:
        mapa, colunas = None, {}
        for linha in aba.iter_rows(values_only=True):
            if mapa is None:
                achado = {COLUNAS[_chave(v)]: i for i, v in enumerate(linha)
                          if _chave(v) in COLUNAS}
                if "ean" in achado and "vum" in achado:
                    mapa, colunas = achado, achado
                continue
            ean = normalizar_ean(linha[colunas["ean"]])
            vum = _numero(linha[colunas["vum"]])
            if not ean or not vum or vum <= 0:
                continue
            # EAN repetido entre abas (raro): fica o de maior giro, que e' o
            # primeiro -- as abas ja vem ordenadas por valor.
            dados.setdefault(ean, {
                "vum": vum,
                "segmento": aba.title.strip().upper(),
                "curva": str(linha[colunas["curva"]]).strip() if "curva" in colunas
                         and linha[colunas["curva"]] else "",
                "posicao": str(linha[colunas["posicao"]]).strip() if "posicao" in colunas
                           and linha[colunas["posicao"]] else "",
            })
    return dados


def atualizar(xlsx: Path, referencia: Path, dry_run: bool = False) -> dict:
    brick = ler_brick(xlsx)
    with referencia.open(encoding="utf-8-sig", newline="") as fh:
        leitor = csv.DictReader(fh)
        campos = leitor.fieldnames
        linhas = list(leitor)

    resumo = {"linhas": len(linhas), "brick_planilha": len(brick),
              "casaram": 0, "novos": 0, "mudaram": 0, "perderam": 0}
    for linha in linhas:
        atual = linha.get("vum_brick", "").strip()
        novo = brick.get(normalizar_ean(linha.get("ean")))
        if novo is None:
            # O Brick anterior fica de fora da planilha nova: o item saiu do
            # painel (ou parou de girar). Preco velho vira preco errado.
            if atual:
                resumo["perderam"] += 1
            linha["vum_brick"] = linha["curva_abc"] = linha["segmento_brick"] = ""
            continue
        resumo["casaram"] += 1
        if not atual:
            resumo["novos"] += 1
        elif f"{float(atual):.4f}" != f"{novo['vum']:.4f}":
            resumo["mudaram"] += 1
        linha["vum_brick"] = f"{novo['vum']:.4f}"
        linha["curva_abc"] = novo["curva"]
        linha["segmento_brick"] = novo["segmento"]

    if dry_run:
        return resumo
    backup = referencia.with_name(
        f"{referencia.stem}_backup_{datetime.now():%Y%m%d_%H%M%S}_pre_brick.csv")
    shutil.copy2(referencia, backup)
    with referencia.open("w", encoding="utf-8-sig", newline="") as fh:
        escritor = csv.DictWriter(fh, fieldnames=campos)
        escritor.writeheader()
        escritor.writerows(linhas)
    resumo["backup"] = backup.name
    return resumo


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("xlsx", type=Path, nargs="?", default=BRICK_ATUAL)
    parser.add_argument("--referencia", type=Path, default=REFERENCIA)
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()
    resumo = atualizar(args.xlsx, args.referencia, args.dry_run)
    for chave, valor in resumo.items():
        print(f"{chave}: {valor}")


if __name__ == "__main__":
    main()
