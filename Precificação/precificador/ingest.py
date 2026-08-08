"""Carregadores idempotentes das fontes de dados para o SQLite do precificador.

Cada carregar_* apaga e recarrega sua tabela (idempotente por fonte) e
devolve o numero de linhas gravadas. main() executa todas em ordem e
registra cada carga em carga_log.
"""
from __future__ import annotations

import csv
import sqlite3
from pathlib import Path

from openpyxl import load_workbook

import db
from caminhos import CONSULTA_PRECOS, MARCA_EXCLUSIVA_XLSX, SUBCATEGORIA_XLSX

ROOT = Path(__file__).parent.parent

CUSTO_NF_XLSX = ROOT / "Relatório notas fiscais 24-07_com_custo_unitario.xlsx"
BRICK_ESTOQUE_XLSX = ROOT / "outputs" / "consolidado_estoque" / "estoque_pmc_brick.xlsx"
PMC_PR_XLSX = ROOT / "outputs" / "eans_pmc" / "ean_descricao_fabricante_pmc_pr.xlsx"
POLITICA_CSV = ROOT / "POLITICA_MARKUP_POR_CATEGORIA.csv"
PRECOS_CSV = CONSULTA_PRECOS / "precos.csv"
EANS_NEGATIVOS_CSV = CONSULTA_PRECOS / "eans_negativos.csv"


def normalizar_ean(valor) -> str:
    return "".join(c for c in str(valor or "") if c.isdigit())


def _numero(valor) -> float | None:
    if isinstance(valor, (int, float)):
        return float(valor)
    if isinstance(valor, str) and valor.strip():
        try:
            return float(valor.strip().replace(",", "."))
        except ValueError:
            return None
    return None


def carregar_custos_nf(conn: sqlite3.Connection) -> int:
    ws = load_workbook(CUSTO_NF_XLSX, read_only=True, data_only=True).active
    linhas = []
    produtos: dict[str, tuple[str, str | None, str | None]] = {}
    for row in ws.iter_rows(min_row=5, values_only=True):
        ean = normalizar_ean(row[9])
        descricao = row[10]
        if not ean or not descricao:
            continue
        quantidade, custo_unitario, valor_total = _numero(row[13]), _numero(row[25]), _numero(row[26])
        if quantidade is None or custo_unitario is None or valor_total is None:
            continue
        if not (quantidade > 0 and custo_unitario > 0 and valor_total > 0):
            continue
        tem_icms_st = 1 if (_numero(row[18]) or 0) > 0 else 0
        linhas.append((ean, quantidade, custo_unitario, valor_total, tem_icms_st))
        grupo_filho = str(row[27]).strip() if row[27] else None
        grupo_pai = str(row[28]).strip() if row[28] else None
        produtos.setdefault(ean, (str(descricao).strip(), grupo_pai, grupo_filho))

    conn.execute("DELETE FROM custo_nf")
    conn.executemany(
        "INSERT INTO custo_nf (ean, quantidade, custo_unitario, valor_total, tem_icms_st) VALUES (?, ?, ?, ?, ?)",
        linhas,
    )
    for ean, (descricao, grupo_pai, grupo_filho) in produtos.items():
        conn.execute(
            "INSERT INTO produto (ean, descricao, grupo_pai_nf, grupo_filho_nf) VALUES (?, ?, ?, ?) "
            "ON CONFLICT(ean) DO UPDATE SET "
            "descricao = COALESCE(produto.descricao, excluded.descricao), "
            "grupo_pai_nf = COALESCE(produto.grupo_pai_nf, excluded.grupo_pai_nf), "
            "grupo_filho_nf = COALESCE(produto.grupo_filho_nf, excluded.grupo_filho_nf)",
            (ean, descricao, grupo_pai, grupo_filho),
        )
    conn.commit()
    return len(linhas)


def carregar_concorrentes(conn: sqlite3.Connection) -> int:
    linhas = []
    with PRECOS_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            ean = normalizar_ean(row.get("ean"))
            if not ean:
                continue
            preco = _numero(row.get("preco"))
            linhas.append((
                ean,
                row.get("site") or "",
                row.get("data_hora") or None,
                row.get("status") or "",
                preco,
                row.get("observacoes") or None,
            ))

    conn.execute("DELETE FROM preco_concorrente")
    conn.executemany(
        "INSERT INTO preco_concorrente (ean, site, data_hora, status, preco, observacoes) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        linhas,
    )
    conn.commit()
    return len(linhas)


def carregar_brick_e_estoque(conn: sqlite3.Connection) -> tuple[int, int]:
    wb = load_workbook(BRICK_ESTOQUE_XLSX, read_only=True, data_only=True)

    segmento_por_ean: dict[str, str] = {}
    totais_por_ean: dict[str, tuple] = {}
    for row in wb["Base Estoque"].iter_rows(min_row=2, values_only=True):
        ean = normalizar_ean(row[0])
        if not ean:
            continue
        if row[9]:
            segmento_por_ean[ean] = str(row[9]).strip()
        totais_por_ean[ean] = (_numero(row[2]), _numero(row[3]), _numero(row[4]))

    brick_linhas, estoque_linhas, produtos = [], [], {}
    for row in wb["Tabela"].iter_rows(min_row=2, values_only=True):
        ean = normalizar_ean(row[0])
        if not ean:
            continue
        descricao = row[1]
        custo_unit, venda_unit, pmc, curva, vum, posicao = row[2], row[3], row[4], row[6], row[7], row[8]

        if _numero(vum):
            brick_linhas.append((
                ean, _numero(vum), curva, str(posicao) if posicao is not None else None,
                segmento_por_ean.get(ean), _numero(pmc),
            ))
        if _numero(custo_unit) or _numero(venda_unit):
            estoque_atual, valor_total_custo, valor_total_venda = totais_por_ean.get(ean, (None, None, None))
            estoque_linhas.append((
                ean, estoque_atual, _numero(custo_unit), _numero(venda_unit),
                valor_total_custo, valor_total_venda,
            ))
        if descricao:
            produtos.setdefault(ean, str(descricao).strip())

    conn.execute("DELETE FROM preco_brick")
    conn.executemany(
        "INSERT INTO preco_brick (ean, vum, curva_abc, posicao_mais_vendidos, segmento, pmc_maximo) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        brick_linhas,
    )
    conn.execute("DELETE FROM estoque")
    conn.executemany(
        "INSERT INTO estoque (ean, estoque_atual, custo_unitario, preco_venda_atual, valor_total_custo, valor_total_venda) "
        "VALUES (?, ?, ?, ?, ?, ?)",
        estoque_linhas,
    )
    for ean, descricao in produtos.items():
        conn.execute(
            "INSERT INTO produto (ean, descricao) VALUES (?, ?) "
            "ON CONFLICT(ean) DO UPDATE SET descricao = excluded.descricao WHERE produto.descricao IS NULL",
            (ean, descricao),
        )
    conn.commit()
    return len(brick_linhas), len(estoque_linhas)


def carregar_pmc_cmed(conn: sqlite3.Connection) -> int:
    # Fonte tem EANs duplicados (mesmo produto de fabricantes/apresentacoes
    # cadastradas mais de uma vez). Convencao documentada na planilha de
    # consolidacao: para EAN duplicado, usar o menor PMC.
    ws = load_workbook(PMC_PR_XLSX, read_only=True, data_only=True)["Produtos"]
    por_ean: dict[str, tuple] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        ean = normalizar_ean(row[0])
        pmc = _numero(row[3])
        if not ean or pmc is None:
            continue
        candidato = (
            ean,
            str(row[1]).strip() if row[1] else None,
            str(row[2]).strip() if row[2] else None,
            pmc,
        )
        atual = por_ean.get(ean)
        if atual is None or pmc < atual[3]:
            por_ean[ean] = candidato

    conn.execute("DELETE FROM pmc_cmed_pr")
    conn.executemany(
        "INSERT INTO pmc_cmed_pr (ean, descricao, fabricante, pmc) VALUES (?, ?, ?, ?)",
        list(por_ean.values()),
    )
    conn.commit()
    return len(por_ean)


def carregar_politica(conn: sqlite3.Connection) -> int:
    linhas = []
    with POLITICA_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f, delimiter=";"):
            try:
                linhas.append((
                    row["categoria_exata"],
                    row["papel"],
                    float(row["lucro_liquido_alvo_pct"].replace(",", ".")) / 100,
                    float(row["fator_fisico_sobre_mediana_web"].replace(",", ".")),
                    row.get("status_observacao"),
                ))
            except (KeyError, ValueError, AttributeError):
                continue

    conn.execute("DELETE FROM politica_categoria")
    conn.executemany(
        "INSERT INTO politica_categoria "
        "(classificacao_exata, papel, lucro_liquido_alvo_pct, fator_fisico_antigo, status_observacao) "
        "VALUES (?, ?, ?, ?, ?)",
        linhas,
    )
    conn.commit()
    return len(linhas)


def carregar_marca_propria(conn: sqlite3.Connection) -> int:
    """Marca produtos de marca propria a partir de eans_negativos.csv do app de
    coleta: lista de EANs que o usuario mantem manualmente para "nunca
    pesquisar" -- confirmado pelo usuario como sendo exatamente a marca
    propria (nao ha concorrente para pesquisar). Validado: os 69 EANs antes
    identificados pelo relatorio de NF de PERFUMARIA sao um subconjunto exato
    dos 216 EANs desta lista.
    """
    eans = set()
    with EANS_NEGATIVOS_CSV.open(encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            codigo = normalizar_ean(row.get("ean"))
            if codigo:
                eans.add(codigo)

    conn.execute("UPDATE produto SET marca_propria = 0")
    conn.executemany(
        "UPDATE produto SET marca_propria = 1 WHERE ean = ?", [(e,) for e in eans]
    )
    conn.commit()
    return len(eans)


def carregar_subcategoria_classificada(conn: sqlite3.Connection) -> int:
    """Subcategoria real por EAN (planilha fornecida pelo usuario: EAN, Descricao,
    Categoria). Os nomes de categoria batem exatamente com classificacao_exata
    da politica (confirmado: PERFUMARIA > ... e EXCLUSIVOS > GERAL). Usada em
    rodada.py para substituir o de-para provisorio (so macro) quando disponivel.
    """
    if not SUBCATEGORIA_XLSX.exists():
        return 0
    ws = load_workbook(SUBCATEGORIA_XLSX, read_only=True, data_only=True)["Plan1"]
    linhas = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        ean = normalizar_ean(row[0])
        categoria = row[2]
        if ean and categoria:
            linhas.append((ean, str(categoria).strip(), SUBCATEGORIA_XLSX.name))

    conn.execute("DELETE FROM subcategoria_classificada")
    conn.executemany(
        "INSERT INTO subcategoria_classificada (ean, classificacao_exata, fonte) VALUES (?, ?, ?)",
        linhas,
    )
    conn.commit()
    return len(linhas)


def carregar_marca_exclusiva(conn: sqlite3.Connection) -> int:
    """Preco de venda dos itens de Marca Exclusiva Associados (planilha do
    usuario, coluna K = 'Preco Venda'), gravado em produto.marca_exclusiva_preco.
    Preco fixado manualmente pelo usuario -- nunca recalculado pela rodada
    (ver rodada_v2.py). Cria o produto se o EAN ainda nao existir na base
    mestre (item pode nao estar no estoque atual).
    """
    if not MARCA_EXCLUSIVA_XLSX.exists():
        return 0
    ws = load_workbook(MARCA_EXCLUSIVA_XLSX, read_only=True, data_only=True)["LISTA PR TRIMESTRAL"]
    itens = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        ean = normalizar_ean(row[1])
        preco = _numero(row[10])
        descricao = row[4]
        if ean and preco:
            itens[ean] = (descricao, round(preco, 2))

    conn.execute("UPDATE produto SET marca_exclusiva_preco = NULL")
    for ean, (descricao, preco) in itens.items():
        existe = conn.execute("SELECT 1 FROM produto WHERE ean = ?", (ean,)).fetchone()
        if existe:
            conn.execute(
                "UPDATE produto SET marca_exclusiva_preco = ? WHERE ean = ?", (preco, ean)
            )
        else:
            conn.execute(
                "INSERT INTO produto (ean, descricao, marca_exclusiva_preco) VALUES (?, ?, ?)",
                (ean, descricao, preco),
            )
        if not conn.execute("SELECT 1 FROM estoque WHERE ean = ?", (ean,)).fetchone():
            conn.execute(
                "INSERT INTO estoque (ean, estoque_atual, preco_venda_atual) VALUES (?, 0, ?)",
                (ean, preco),
            )
    conn.commit()
    return len(itens)


def main() -> None:
    conn = db.connect()
    db.criar_schema(conn)

    cargas = [
        ("custo_nf", CUSTO_NF_XLSX, lambda: carregar_custos_nf(conn)),
        ("preco_concorrente", PRECOS_CSV, lambda: carregar_concorrentes(conn)),
        ("politica_categoria", POLITICA_CSV, lambda: carregar_politica(conn)),
        ("pmc_cmed_pr", PMC_PR_XLSX, lambda: carregar_pmc_cmed(conn)),
    ]
    for fonte, arquivo, funcao in cargas:
        n = funcao()
        db.registrar_carga(conn, fonte, arquivo.name, n)
        print(f"{fonte:20s} {n:6d} linhas  <- {arquivo.name}")

    n_brick, n_estoque = carregar_brick_e_estoque(conn)
    db.registrar_carga(conn, "preco_brick", BRICK_ESTOQUE_XLSX.name, n_brick)
    db.registrar_carga(conn, "estoque", BRICK_ESTOQUE_XLSX.name, n_estoque)
    print(f"{'preco_brick':20s} {n_brick:6d} linhas  <- {BRICK_ESTOQUE_XLSX.name}")
    print(f"{'estoque':20s} {n_estoque:6d} linhas  <- {BRICK_ESTOQUE_XLSX.name}")

    n_marca_propria = carregar_marca_propria(conn)
    db.registrar_carga(conn, "marca_propria", EANS_NEGATIVOS_CSV.name, n_marca_propria)
    print(f"{'marca_propria':20s} {n_marca_propria:6d} EANs marcados")

    n_subcategoria = carregar_subcategoria_classificada(conn)
    db.registrar_carga(conn, "subcategoria_classificada", SUBCATEGORIA_XLSX.name, n_subcategoria)
    print(f"{'subcategoria':20s} {n_subcategoria:6d} EANs classificados")

    n_marca_exclusiva = carregar_marca_exclusiva(conn)
    db.registrar_carga(conn, "marca_exclusiva", MARCA_EXCLUSIVA_XLSX.name, n_marca_exclusiva)
    print(f"{'marca_exclusiva':20s} {n_marca_exclusiva:6d} EANs com preco manual")

    n_produto = conn.execute("SELECT COUNT(*) FROM produto").fetchone()[0]
    print(f"{'produto (total)':20s} {n_produto:6d} EANs distintos")
    conn.close()


if __name__ == "__main__":
    main()
