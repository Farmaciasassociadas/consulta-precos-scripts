"""Fase 5: exporta a rodada mais recente para Excel, no formato ja usado
pela equipe (mesma logica de abas/formatacao de gerar_precificacao.py),
com as colunas novas desta revisao (Brick, natureza fiscal, divergencia).
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import db

SAIDA_DIR = Path(__file__).parent.parent
COLUNAS_MOEDA = {
    "Custo médio", "Piso", "Alvo/Máx. recomendado", "Mediana/Referência mercado",
    "VUM Brick", "Teto CMED", "Preço atual", "Preço sugerido",
}

CABECALHO = [
    "EAN", "Descrição", "Categoria (provisória)", "Natureza fiscal", "Tier", "Status",
    "Custo médio", "Compras NF", "Piso", "Alvo/Máx. recomendado",
    "Mediana/Referência mercado", "Nº concorrentes", "CV", "Peso Brick", "VUM Brick",
    "Teto CMED", "Preço atual", "Preço sugerido", "Justificativa",
]


def _linha(r: sqlite3.Row) -> list:
    return [
        r["ean"], r["descricao"], r["categoria_provisoria"], r["natureza_fiscal"], r["tier"], r["status"],
        r["custo"], r["n_compras_nf"], r["piso"], r["alvo"],
        r["mercado_referencia"], r["n_concorrentes"], r["cv"], r["peso_brick"], r["vum_brick"],
        r["teto_cmed"], r["preco_atual"], r["preco_sugerido"], r["justificativa"],
    ]


def _formatar(sheet) -> None:
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for coluna in sheet.columns:
        letra = coluna[0].column_letter
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=8)
        sheet.column_dimensions[letra].width = min(largura + 2, 52)
    cabecalho_valores = [c.value for c in sheet[1]]
    for indice, nome in enumerate(cabecalho_valores, start=1):
        if nome in COLUNAS_MOEDA:
            for cell in sheet.iter_rows(min_row=2, min_col=indice, max_col=indice):
                cell[0].number_format = "R$ #,##0.00"


def exportar(conn: sqlite3.Connection, rodada_id: int) -> Path:
    conn.row_factory = sqlite3.Row
    linhas = conn.execute(
        "SELECT * FROM recomendacao WHERE rodada_id = ? ORDER BY categoria_provisoria, ean", (rodada_id,)
    ).fetchall()

    wb = Workbook()

    completa = wb.active
    completa.title = "Precificação Completa"
    completa.append(CABECALHO)
    for r in linhas:
        completa.append(_linha(r))

    revisao = wb.create_sheet("Revisão Manual")
    revisao.append(CABECALHO)
    for r in linhas:
        if r["status"] != "OK":
            revisao.append(_linha(r))

    piso_hoje = wb.create_sheet("Abaixo do Piso Hoje")
    piso_hoje.append(CABECALHO)
    for r in linhas:
        if r["piso"] is not None and r["preco_atual"] is not None and 0 < r["preco_atual"] < r["piso"]:
            piso_hoje.append(_linha(r))

    divergencia = wb.create_sheet("Divergência Brick x Web")
    divergencia.append(CABECALHO)
    for r in linhas:
        if r["status"] == "DIVERGENCIA_BRICK_WEB":
            divergencia.append(_linha(r))

    resumo = wb.create_sheet("Resumo", 0)
    resumo.append(["Status", "Itens"])
    contagem_status: dict[str, int] = {}
    for r in linhas:
        contagem_status[r["status"]] = contagem_status.get(r["status"], 0) + 1
    for status, n in sorted(contagem_status.items(), key=lambda x: -x[1]):
        resumo.append([status, n])
    resumo.append([])
    resumo.append(["Categoria (provisória)", "Itens", "OK", "Preço médio sugerido"])
    por_categoria: dict[str, list] = {}
    for r in linhas:
        chave = r["categoria_provisoria"] or "(sem categoria)"
        por_categoria.setdefault(chave, []).append(r)
    for categoria, rows in sorted(por_categoria.items()):
        ok = [r["preco_sugerido"] for r in rows if r["status"] == "OK" and r["preco_sugerido"] is not None]
        media = round(sum(ok) / len(ok), 2) if ok else None
        resumo.append([categoria, len(rows), len(ok), media])

    notas = wb.create_sheet("Notas e Premissas")
    for linha in [
        "Rodada gerada por precificador/rodada.py sobre precificador.db (Fase 1-4 do PLANO_SISTEMA_PRECIFICACAO.md).",
        "Custo: media do custo unitario validado por NF (Valor Total do Item / Qtde. Unitaria), nao e custo de reposicao.",
        "Piso: max(custo / divisor fiscal segregado, custo + contribuicao minima por unidade).",
        "Divisor fiscal segregado por natureza (medicamento/perfumaria_higiene/padrao) -- ver plano Parte 2.",
        "Mercado: blend Brick (VUM, com spread de etiqueta) + mediana web filtrada em 4 camadas de outlier, peso do Brick variavel conforme forca da coleta web.",
        "Categoria (provisoria): de-para PROVISORIO apenas no nivel macro da taxonomia antiga -> nova; SIMILAR e LIBERADO ainda sem politica utilizavel (ver rodada.py).",
        "Teto CMED: aplicado quando o EAN tem PMC-PR cadastrado; grade de arredondamento nunca ultrapassa o teto.",
        "Trava de variacao: preco so e sugerido automaticamente se a mudanca vs. preco praticado hoje for <= 15%; acima disso, vai para revisao manual.",
        "Nenhum preco desta planilha deve ser aplicado sem conferencia humana das linhas em revisao manual.",
        "Descricao terminada em ' *': marca propria (fonte: eans_negativos.csv do app de coleta -- "
        "lista de 'nunca pesquisar' mantida manualmente, confirmada pelo usuario como marca propria). "
        "Preco fica fora da precificacao automatica -- inserir manualmente.",
    ]:
        notas.append([linha])

    for sheet in wb.worksheets:
        if sheet.title not in ("Resumo", "Notas e Premissas"):
            _formatar(sheet)

    caminho = SAIDA_DIR / f"PRECIFICACAO_RODADA_{rodada_id}.xlsx"
    wb.save(caminho)
    return caminho


def main() -> None:
    conn = db.connect()
    rodada_id = conn.execute("SELECT id FROM rodada ORDER BY id DESC LIMIT 1").fetchone()[0]
    caminho = exportar(conn, rodada_id)
    conn.close()
    print(f"Exportado: {caminho}")


if __name__ == "__main__":
    main()
