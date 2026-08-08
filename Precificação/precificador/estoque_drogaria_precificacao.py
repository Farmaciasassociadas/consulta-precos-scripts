"""Cruza o relatorio de estoque da DROGARIA (Google Drive) com a base mestre
de precificacao (precificador.db) e gera uma planilha com preco sugerido,
Brick, precos coletados por farmacia, observacoes e curva ABC.

Nao grava nada em precificador.db (nao cria rodada oficial): saida isolada
em Excel, para conferencia manual antes de qualquer aplicacao de preco.
"""
from __future__ import annotations

import difflib
import re
import sqlite3
import unicodedata
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from statistics import median

import openpyxl
from openpyxl.styles import Font, PatternFill

import db
from caminhos import ESTOQUE_XLSX
from engine import economico, mercado, parametros
SAIDA_XLSX = Path(__file__).parent.parent / "ESTOQUE_DROGARIA_PRECIFICADO.xlsx"

SITES = [
    "drogaraia", "nissei", "saopaulo", "saojoao", "panvel",
    "sistema", "paguemenos", "precopopular", "drogariasp", "farmasp",
]

DEPARA_PROVISORIO = {
    "GENÉRICO": "GENERICO", "GENERICO": "GENERICO",
    "SIMILAR": "SIMILAR",
    "REFERÊNCIA": "ETICOS", "REFERENCIA": "ETICOS",
    "PERFUMARIA": "PERFUMARIA",
}

SIMILARIDADE_MINIMA = 0.92

# Razoes de embalagem plausiveis: quantidades comuns em caixas (uni, pack de 2-3,
# blister de 4-12, cartelas de 24-30, multipacks de 48-100). Se nenhuma razao
# fixa funcionar, a funcao tenta todos os inteiros 2..100.
RAZOES_EMBALAGEM_FIXAS = [2, 3, 4, 5, 6, 8, 10, 12, 15, 20, 24, 25, 30, 40, 48, 50, 60, 100]


def corrigir_preco_atual(
    venda_unit: float | None,
    preco_anterior: float | None,
    custo_final: float | None,
    pmc_val: float | None,
    vum_brick: float | None,
) -> tuple[float | None, str | None]:
    """Detecta e corrige (ou descarta) um preco unitario do estoque.xlsx
    absurdo. PMC/CMED nao serve como ancora unica: o cadastro oficial pode
    estar preso a uma apresentacao (qtde de unidades) diferente da caixa
    comercial, o que faz o PMC concordar com o valor errado (caso observado:
    EAN 7896009498411, PMC registrado para caixa X60, produto real e C/3).
    A ancora principal segue a confiabilidade observada no pipeline:
    Brick > PMC > preco anterior > custo.
    Devolve (valor_corrigido_ou_None, nota_para_observacao)."""
    if venda_unit is None:
        return None, None

    ancoras = []
    # Brick primeiro (mais confiavel, vem do fabricante).
    if vum_brick:
        ancoras.append(("Brick", vum_brick, 2.0))
    # PMC segundo (regulatorio, mas pode estar para apresentacao diferente).
    if pmc_val:
        ancoras.append(("PMC/CMED", pmc_val, 1.05))
    # Preco anterior terceiro (pode ter sido errado da coleta anterior).
    if preco_anterior:
        ancoras.append(("preco anterior no banco", preco_anterior, 3.0))
    # Custo por ultimo (eh o alvo economico minimo, nao a unidade de venda).
    if custo_final:
        ancoras.append(("custo (markup maximo 10x)", custo_final, 10.0))
    if not ancoras:
        return venda_unit, None

    nome_ref, ancora, tolerancia = ancoras[0]
    if venda_unit <= ancora * tolerancia:
        return venda_unit, None

    # Tenta razoes fixas contra a principal
    melhor_razao, melhor_dif, melhor_ref = None, None, None
    for razao in RAZOES_EMBALAGEM_FIXAS:
        corrigido = venda_unit / razao
        if ancora * 0.2 <= corrigido <= ancora * tolerancia:
            dif = abs(corrigido - ancora)
            if melhor_dif is None or dif < melhor_dif:
                melhor_razao, melhor_dif, melhor_ref = razao, dif, nome_ref

    # Se razoes fixas nao acharam: tenta inteiros 2..100 contra a principal,
    # e se ainda nao achar, tenta contra as outras ancoras (PMC, Brick, custo)
    if not melhor_razao:
        razao_bruta = venda_unit / ancora
        banda_min = max(0.10, 1.5 / razao_bruta) if razao_bruta > 0 else 0.10
        for razao in range(2, 101):
            if razao in RAZOES_EMBALAGEM_FIXAS:
                continue
            corrigido = venda_unit / razao
            if ancora * banda_min <= corrigido <= ancora * tolerancia:
                dif = abs(corrigido - ancora)
                if melhor_dif is None or dif < melhor_dif:
                    melhor_razao, melhor_dif, melhor_ref = razao, dif, nome_ref

    # Se ainda nao achou: tenta contra qualquer outra ancora disponivel.
    # Ordem: PMC > Brick > custo (deixa preco anterior por ultimo).
    if not melhor_razao:
        outras_ancoras = []
        if pmc_val and pmc_val != ancora:
            outras_ancoras.append(("PMC/CMED", pmc_val, 1.05))
        if vum_brick and vum_brick != ancora:
            outras_ancoras.append(("Brick", vum_brick, 2.0))
        if custo_final and custo_final != ancora:
            outras_ancoras.append(("custo", custo_final * 10, 10.0))
        for ref_nome, ref_valor, ref_tol in outras_ancoras:
            for razao in range(2, 101):
                corrigido = venda_unit / razao
                razao_bruta = venda_unit / ref_valor
                banda_min = max(0.10, 1.5 / razao_bruta) if razao_bruta > 0 else 0.10
                if ref_valor * banda_min <= corrigido <= ref_valor * ref_tol:
                    dif = abs(corrigido - ref_valor)
                    if melhor_dif is None or dif < melhor_dif:
                        melhor_razao, melhor_dif, melhor_ref = razao, dif, ref_nome

    if melhor_razao:
        corrigido = venda_unit / melhor_razao
        return corrigido, (
            f"preco do estoque.xlsx incoerente com {melhor_ref} (R$ {venda_unit:.2f} vs R$ {ancora:.2f}); "
            f"corrigido dividindo por {melhor_razao} (provavel erro de embalagem/preco no ERP)."
        )

    return None, (
        f"preco do estoque.xlsx descartado: R$ {venda_unit:.2f} incoerente com {nome_ref} "
        f"(R$ {ancora:.2f}) e nenhuma razao de embalagem ate 100x explicou a diferenca."
    )


def _normalizar(texto: str | None) -> str:
    if not texto:
        return ""
    texto = unicodedata.normalize("NFKD", texto).encode("ascii", "ignore").decode("ascii")
    texto = texto.upper()
    texto = re.sub(r"[^A-Z0-9 ]", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


@dataclass
class ItemEstoque:
    ean: str
    descricao: str
    grupo_pai: str | None
    grupo_filho: str | None
    estoque_atual: float
    valor_total_custo: float
    valor_total_venda: float

    @property
    def custo_unit(self) -> float | None:
        if self.estoque_atual and self.valor_total_custo:
            return self.valor_total_custo / self.estoque_atual
        return None

    @property
    def venda_unit(self) -> float | None:
        if self.estoque_atual and self.valor_total_venda:
            return self.valor_total_venda / self.estoque_atual
        return None


def ler_estoque(caminho: Path) -> list[ItemEstoque]:
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
    ws = wb.active
    itens = []
    for r in ws.iter_rows(values_only=True):
        if len(r) < 17:
            continue
        ean = r[8]
        descricao = r[9]
        if not ean or not isinstance(ean, str) or not ean.strip().isdigit():
            continue
        if not descricao:
            continue
        estoque_atual = r[12] or 0
        valor_total_custo = r[15] or 0
        valor_total_venda = r[16] or 0
        itens.append(ItemEstoque(
            ean=ean.strip(), descricao=str(descricao).strip(),
            grupo_pai=(r[4] or None), grupo_filho=(r[5] or None),
            estoque_atual=float(estoque_atual),
            valor_total_custo=float(valor_total_custo),
            valor_total_venda=float(valor_total_venda),
        ))
    wb.close()
    return itens


def carregar_base_mestre(conn: sqlite3.Connection):
    conn.row_factory = sqlite3.Row

    produtos = {r["ean"]: r for r in conn.execute("SELECT * FROM produto")}

    custo_nf: dict[str, dict] = {}
    for r in conn.execute(
        "SELECT ean, AVG(custo_unitario) AS custo_medio, COUNT(*) AS n, MAX(tem_icms_st) AS icms "
        "FROM custo_nf GROUP BY ean"
    ):
        custo_nf[r["ean"]] = {"custo": r["custo_medio"], "n": r["n"], "icms": bool(r["icms"])}

    brick = {r["ean"]: r for r in conn.execute("SELECT * FROM preco_brick")}
    pmc = {r["ean"]: r["pmc"] for r in conn.execute("SELECT ean, pmc FROM pmc_cmed_pr") if r["pmc"]}
    subcategoria = {
        r["ean"]: r["classificacao_exata"]
        for r in conn.execute("SELECT ean, classificacao_exata FROM subcategoria_classificada")
    }
    politica = {
        r["classificacao_exata"]: (r["papel"], r["lucro_liquido_alvo_pct"])
        for r in conn.execute("SELECT * FROM politica_categoria")
    }

    eans_com_coleta = {r[0] for r in conn.execute("SELECT DISTINCT ean FROM preco_concorrente")}

    preco_anterior = {
        r["ean"]: r["preco_venda_atual"]
        for r in conn.execute("SELECT ean, preco_venda_atual FROM estoque")
        if r["preco_venda_atual"]
    }

    historico_por_ean: dict[str, list[float]] = {}
    for r in conn.execute("SELECT ean, preco_sugerido FROM recomendacao WHERE preco_sugerido IS NOT NULL"):
        historico_por_ean.setdefault(r["ean"], []).append(r["preco_sugerido"])
    mediana_historico = {ean: median(valores) for ean, valores in historico_por_ean.items()}

    return (
        produtos, custo_nf, brick, pmc, subcategoria, politica, eans_com_coleta, preco_anterior,
        mediana_historico,
    )


def observacoes_do_ean(conn: sqlite3.Connection, ean: str) -> list[mercado.Observacao]:
    linhas = conn.execute(
        "SELECT site, data_hora, status, preco, observacoes FROM preco_concorrente WHERE ean = ?", (ean,)
    ).fetchall()
    return [
        mercado.Observacao(
            site=site, preco=preco, status=status,
            data_hora=mercado.parse_data_hora(data_hora), observacoes=observacoes,
        )
        for site, data_hora, status, preco, observacoes in linhas
    ]


def precos_por_site(conn: sqlite3.Connection, ean: str) -> dict[str, float | None]:
    linhas = conn.execute(
        "SELECT site, data_hora, status, preco FROM preco_concorrente WHERE ean = ? ORDER BY data_hora",
        (ean,),
    ).fetchall()
    resultado: dict[str, float | None] = {}
    for site, data_hora, status, preco in linhas:
        if status == "OK" and preco:
            resultado[site] = preco  # ultima OK (linhas em ordem por data_hora)
    return resultado


def montar_indice_nomes(produtos: dict) -> list[tuple[str, str]]:
    return [(ean, _normalizar(row["descricao"])) for ean, row in produtos.items() if row["descricao"]]


def buscar_por_nome(descricao: str, indice: list[tuple[str, str]]) -> tuple[str | None, float]:
    alvo = _normalizar(descricao)
    melhor_ean, melhor_score = None, 0.0
    for ean, nome in indice:
        score = difflib.SequenceMatcher(None, alvo, nome).ratio()
        if score > melhor_score:
            melhor_ean, melhor_score = ean, score
    return melhor_ean, melhor_score


def main() -> None:
    itens = ler_estoque(ESTOQUE_XLSX)
    print(f"Itens lidos do estoque.xlsx: {len(itens)}")

    conn = db.connect()
    (
        produtos, custo_nf, brick, pmc, subcategoria, politica, eans_com_coleta, preco_anterior,
        mediana_historico,
    ) = carregar_base_mestre(conn)
    indice_nomes = montar_indice_nomes(produtos)

    lucros_validos = [v[1] for v in politica.values() if v[1] is not None]
    lucro_mediano_fallback = median(lucros_validos) if lucros_validos else 0.15

    params = parametros.carregar()
    hoje = date.today()

    linhas_saida = []
    for item in itens:
        obs_extra: list[str] = []

        # --- casamento por EAN, depois por nome ---
        ean_match = None
        match_tipo = None
        if item.ean in produtos or item.ean in custo_nf or item.ean in brick or item.ean in pmc \
                or item.ean in subcategoria or item.ean in eans_com_coleta:
            ean_match = item.ean
            match_tipo = "EAN"
        else:
            candidato, score = buscar_por_nome(item.descricao, indice_nomes)
            if candidato and score >= SIMILARIDADE_MINIMA:
                ean_match = candidato
                match_tipo = "NOME"
                obs_extra.append(f"match por nome (similaridade {score:.2f}), nao por EAN direto -- conferir.")
            else:
                obs_extra.append("EAN nao encontrado na base mestre; sem correspondencia proxima por nome tambem.")

        preco_exclusivo = produtos.get(ean_match)["marca_exclusiva_preco"] if ean_match and produtos.get(ean_match) else None
        if preco_exclusivo is not None:
            linhas_saida.append({
                "ean": item.ean,
                "descricao": item.descricao,
                "custo_unitario": None,
                "preco_venda_atual": item.venda_unit,
                "preco_sugerido": preco_exclusivo,
                "preco_brick": None,
                "sites": {},
                "observacoes": (
                    "Marca Exclusiva Associados: preco de venda fixado manualmente pelo usuario "
                    "(planilha Marca Exclusiva Associados TRATADO.xlsx, coluna Preco Venda), "
                    "fora da precificacao automatica."
                ),
                "curva_abc": None,
                "status": "MARCA_EXCLUSIVA_MANUAL",
                "tier": None,
                "match_tipo": match_tipo or "SEM_MATCH",
            })
            continue

        if ean_match and produtos.get(ean_match) and produtos[ean_match]["marca_propria"]:
            linhas_saida.append({
                "ean": item.ean,
                "descricao": f"{item.descricao} *",
                "custo_unitario": None,
                "preco_venda_atual": item.venda_unit,
                "preco_sugerido": None,
                "preco_brick": None,
                "sites": {},
                "observacoes": "Marca propria: preco definido manualmente pelo usuario, fora da precificacao automatica.",
                "curva_abc": None,
                "status": "MARCA_PROPRIA_MANUAL",
                "tier": None,
                "match_tipo": match_tipo or "SEM_MATCH",
            })
            continue

        # --- custo ---
        custo_final = None
        fonte_custo = None
        tem_icms_st = False
        if ean_match and ean_match in custo_nf:
            custo_final = custo_nf[ean_match]["custo"]
            tem_icms_st = custo_nf[ean_match]["icms"]
            fonte_custo = "NF"
        elif item.custo_unit and item.custo_unit > 0:
            custo_final = item.custo_unit
            fonte_custo = "ESTOQUE"
            obs_extra.append("custo obtido do estoque.xlsx (Valor Total de Custo / Estoque Atual), nao validado por NF.")
        else:
            obs_extra.append("sem custo em nenhuma fonte (NF ou estoque).")

        # --- brick / pmc ---
        brick_row = brick.get(ean_match) if ean_match else None
        vum_brick = brick_row["vum"] if brick_row else None
        curva_abc = brick_row["curva_abc"] if brick_row else None
        segmento_brick = brick_row["segmento"] if brick_row else None
        pmc_val = pmc.get(ean_match) if ean_match else None

        preco_atual, nota_correcao = corrigir_preco_atual(
            item.venda_unit, preco_anterior.get(ean_match), custo_final, pmc_val, vum_brick
        )
        if nota_correcao:
            obs_extra.append(nota_correcao)

        mediana_hist_ean = mediana_historico.get(ean_match) if ean_match else None
        if economico.preco_atual_e_outlier_historico(preco_atual, mediana_hist_ean):
            obs_extra.append(
                f"preco atual (R$ {preco_atual:.2f}) destoa da mediana historica de sugestoes "
                f"(R$ {mediana_hist_ean:.2f}); cadastro provavelmente desatualizado, ignorado como referencia."
            )

        # --- mercado ---
        obs_concorrentes = observacoes_do_ean(conn, ean_match) if ean_match else []
        resultado_mercado = mercado.calcular_mercado(
            obs_concorrentes, params, hoje, vum_brick=vum_brick, segmento_brick=segmento_brick,
        )

        # --- categoria / politica ---
        classificacao_exata = subcategoria.get(ean_match) if ean_match else None
        if not classificacao_exata and item.grupo_pai and item.grupo_filho:
            candidata = f"{item.grupo_pai} > {item.grupo_filho}"
            if candidata in politica:
                classificacao_exata = candidata
        if not classificacao_exata and item.grupo_pai in politica:
            classificacao_exata = item.grupo_pai
        if not classificacao_exata and ean_match and ean_match in produtos:
            classificacao_exata = DEPARA_PROVISORIO.get((produtos[ean_match]["grupo_pai_nf"] or "").strip())

        papel, lucro_pct = politica.get(classificacao_exata, (None, None)) if classificacao_exata else (None, None)

        natureza = economico.natureza_fiscal(classificacao_exata, tem_icms_st)
        tier = economico.determinar_tier(
            papel_politica=papel, curva_abc=curva_abc,
            n_concorrentes=resultado_mercado.n, cv=resultado_mercado.cv,
            tem_brick=vum_brick is not None,
        )
        if tier == "REVISAO_HUMANA":
            # Controlado/fracionado: o motor oficial bloqueia sugestao por
            # padrao (rodada.py, producao). Aqui habilitamos por excecao com
            # o tier mais conservador (protege margem, nao persegue mercado
            # para baixo) em vez de deixar sem preco algum.
            tier = "PROTECAO_MARGEM"
            obs_extra.append(
                "categoria de revisao humana (controlado/fracionado): sugestao automatica habilitada por "
                "excecao, tier PROTECAO_MARGEM -- conferir apresentacao/dosagem com atencao redobrada."
            )

        resultado = economico.aplicar_travas(
            custo=custo_final, natureza_fiscal_item=natureza, tier=tier,
            valor_referencia_mercado=resultado_mercado.valor_referencia,
            divergencia_brick_web=resultado_mercado.divergencia_brick_web,
            lucro_liquido_alvo_pct=lucro_pct, teto_cmed=pmc_val,
            preco_atual=preco_atual, params=params,
        )

        preco_sugerido = resultado.preco_sugerido
        status = resultado.status
        justificativa = resultado.justificativa

        # --- fallback: custo existe mas sem politica de categoria cadastrada ---
        if preco_sugerido is None and status == "REVISAO_MANUAL_SEM_MARKUP":
            resultado2 = economico.aplicar_travas(
                custo=custo_final, natureza_fiscal_item=natureza, tier=tier,
                valor_referencia_mercado=resultado_mercado.valor_referencia,
                divergencia_brick_web=resultado_mercado.divergencia_brick_web,
                lucro_liquido_alvo_pct=lucro_mediano_fallback, teto_cmed=pmc_val,
                preco_atual=preco_atual, params=params,
            )
            preco_sugerido = resultado2.preco_sugerido
            status = (status + "_MARGEM_ESTIMADA") if resultado2.preco_sugerido is not None else status
            justificativa = resultado2.justificativa
            obs_extra.append(
                f"sem politica de categoria cadastrada para '{classificacao_exata or 'sem classificacao'}': "
                f"aplicada margem padrao mediana ({lucro_mediano_fallback * 100:.0f}%) como estimativa -- "
                "revisar classificacao da categoria."
            )

        if custo_final is None and resultado_mercado.valor_referencia is None:
            obs_extra.append("sem custo e sem nenhuma referencia de mercado (Brick/coleta): preco nao pode ser sugerido, cadastro manual necessario.")

        # --- custo implicito quando o preco veio so do mercado (sem custo real) ---
        custo_exibido = custo_final
        if custo_final is None and preco_sugerido is not None and lucro_pct is not None:
            divisor = economico.divisor_alvo(params, natureza, lucro_pct)
            if divisor > 0:
                custo_exibido = round(preco_sugerido * divisor, 4)
                obs_extra.append(
                    f"custo estimado (implicito) = preco sugerido / markup-alvo da categoria "
                    f"(R$ {custo_exibido:.2f}) -- confirmar custo real assim que possivel."
                )

        # --- PMC ---
        if pmc_val and preco_atual and preco_atual > pmc_val:
            obs_extra.append(f"preco atual praticado (R$ {preco_atual:.2f}) esta acima do PMC/CMED (R$ {pmc_val:.2f}).")
        if pmc_val and preco_sugerido and preco_sugerido > pmc_val:
            obs_extra.append(f"ATENCAO: preco sugerido (R$ {preco_sugerido:.2f}) ultrapassa o teto PMC/CMED (R$ {pmc_val:.2f}).")

        # --- posicao entre concorrentes (2o/3o mais barato ja e bom) ---
        precos_conc_validos = sorted(
            o.preco for o in obs_concorrentes
            if o.status == "OK" and o.preco and not mercado._e_promocional(o.observacoes)
        )
        if precos_conc_validos and preco_sugerido is not None:
            posicao = 1 + sum(1 for p in precos_conc_validos if p < preco_sugerido)
            if posicao > 3:
                if len(precos_conc_validos) >= 3:
                    preco_3o = precos_conc_validos[2]
                    if resultado.piso is not None and preco_3o < resultado.piso:
                        obs_extra.append(
                            f"posicao {posicao}o entre {len(precos_conc_validos)} concorrentes coletados; "
                            f"custo/piso tecnico (R$ {resultado.piso:.2f}) nao permitiu ficar entre os 3 mais baratos "
                            f"(3o colocado R$ {preco_3o:.2f}) -- mantida precificacao normal, sem prejuizo."
                        )
                    else:
                        obs_extra.append(
                            f"posicao {posicao}o entre {len(precos_conc_validos)} concorrentes coletados "
                            "(fora do top 3) -- avaliar se cabe reduzir dentro da politica."
                        )
                else:
                    obs_extra.append(
                        f"posicao {posicao}o entre {len(precos_conc_validos)} concorrentes coletados (poucos dados)."
                    )

        linha = {
            "ean": item.ean,
            "descricao": item.descricao,
            "custo_unitario": custo_exibido,
            "preco_venda_atual": preco_atual,
            "preco_sugerido": preco_sugerido,
            "preco_brick": vum_brick,
            "sites": precos_por_site(conn, ean_match) if ean_match else {},
            "observacoes": " | ".join(obs_extra) if obs_extra else "",
            "curva_abc": curva_abc,
            "status": status,
            "tier": tier,
            "match_tipo": match_tipo or "SEM_MATCH",
        }
        linhas_saida.append(linha)

    eans_do_estoque = {item.ean for item in itens}
    for ean, produto in produtos.items():
        if produto["marca_exclusiva_preco"] is None or ean in eans_do_estoque:
            continue
        linhas_saida.append({
            "ean": ean,
            "descricao": produto["descricao"],
            "custo_unitario": None,
            "preco_venda_atual": None,
            "preco_sugerido": produto["marca_exclusiva_preco"],
            "preco_brick": None,
            "sites": {},
            "observacoes": (
                "Marca Exclusiva Associados: item fora do estoque atual (nao presente em estoque.xlsx), "
                "incluido manualmente com preco de venda fixado pelo usuario "
                "(planilha Marca Exclusiva Associados TRATADO.xlsx, coluna Preco Venda), "
                "fora da precificacao automatica."
            ),
            "curva_abc": None,
            "status": "MARCA_EXCLUSIVA_MANUAL",
            "tier": None,
            "match_tipo": "SEM_ESTOQUE",
        })

    conn.close()
    exportar(linhas_saida)


def exportar(linhas: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Estoque Precificado"

    cabecalho = [
        "EAN", "Descricao", "Custo unitario", "Preco venda atual", "Preco sugerido", "Preco Brick (VUM)",
        *[s.upper() for s in SITES],
        "Observacoes", "Curva ABC", "Status", "Tier", "Match",
    ]
    ws.append(cabecalho)

    moeda_idx = {3, 4, 5, 6} | {7 + i for i in range(len(SITES))}
    for linha in linhas:
        row = [
            linha["ean"], linha["descricao"], linha["custo_unitario"], linha["preco_venda_atual"],
            linha["preco_sugerido"], linha["preco_brick"],
            *[linha["sites"].get(s) for s in SITES],
            linha["observacoes"], linha["curva_abc"], linha["status"], linha["tier"], linha["match_tipo"],
        ]
        ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for idx in moeda_idx:
        for cell in ws.iter_rows(min_row=2, min_col=idx, max_col=idx):
            cell[0].number_format = "R$ #,##0.00"
    for coluna in ws.columns:
        letra = coluna[0].column_letter
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=8)
        ws.column_dimensions[letra].width = min(largura + 2, 60)

    wb.save(SAIDA_XLSX)
    print(f"Exportado: {SAIDA_XLSX}")


if __name__ == "__main__":
    main()
