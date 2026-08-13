"""Lista os EANs que NUNCA foram consultados em site nenhum.

Nao e' "o site nao achou": e' que nenhuma coleta jamais tentou este EAN. Sao
806 no catalogo de 12/08/2026 -- o vao maior da base, e a razao de a fila de
divergencia ser dominada por itens com 1-2 precos.

Filtros, na ordem em que foram pedidos:
  - fora quem NUNCA apareceu no estoque do ERP (111): produto que a loja nunca
    teve nao e' prioridade de pesquisa;
  - marca propria fica na planilha mas em aba separada e marcada -- nao ha o que
    pesquisar (o preco vem da lista do fornecedor), e misturar as duas listas
    faz alguem perder tempo procurando o que nao existe na concorrencia.
"""
from __future__ import annotations

import sqlite3
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

import db

SAIDA = Path(__file__).parent.parent / "eans_nunca_pesquisados.xlsx"
CABECALHO = ["EAN", "Descrição", "Categoria", "Estoque atual",
             "Valor em estoque (venda)", "Preço de venda atual", "Custo médio"]

CONSULTA = """
    SELECT r.ean, r.descricao, r.categoria_provisoria,
           e.estoque_atual, e.valor_total_venda, e.preco_venda_atual, r.custo,
           COALESCE(pr.marca_propria, 0) AS marca_propria
    FROM recomendacao r
    JOIN estoque e ON e.ean = r.ean
    LEFT JOIN produto pr ON pr.ean = r.ean
    WHERE r.rodada_id = ?
      AND NOT EXISTS (SELECT 1 FROM preco_concorrente p WHERE p.ean = r.ean)
    ORDER BY COALESCE(e.valor_total_venda, 0) DESC
"""


def _formatar(aba) -> None:
    aba.freeze_panes = "A2"
    aba.auto_filter.ref = aba.dimensions
    for celula in aba[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor="1F4E78")
    for coluna in aba.columns:
        letra = coluna[0].column_letter
        largura = max((len(str(c.value)) for c in coluna if c.value is not None), default=8)
        aba.column_dimensions[letra].width = min(largura + 2, 60)
    for indice, nome in enumerate([c.value for c in aba[1]], start=1):
        if "R$" in str(nome) or nome in ("Valor em estoque (venda)", "Preço de venda atual", "Custo médio"):
            for linha in aba.iter_rows(min_row=2, min_col=indice, max_col=indice):
                linha[0].number_format = "R$ #,##0.00"


def exportar(conn: sqlite3.Connection, rodada_id: int) -> tuple[Path, int, int]:
    conn.row_factory = sqlite3.Row
    linhas = conn.execute(CONSULTA, (rodada_id,)).fetchall()
    pesquisar = [r for r in linhas if not r["marca_propria"]]
    propria = [r for r in linhas if r["marca_propria"]]

    wb = Workbook()
    aba = wb.active
    aba.title = "Pesquisar"
    aba.append(CABECALHO)
    for r in pesquisar:
        aba.append([r["ean"], r["descricao"], r["categoria_provisoria"], r["estoque_atual"],
                    r["valor_total_venda"], r["preco_venda_atual"], r["custo"]])
    _formatar(aba)

    aba2 = wb.create_sheet("Marca própria (não pesquisar)")
    aba2.append(CABECALHO)
    for r in propria:
        aba2.append([r["ean"], r["descricao"], r["categoria_provisoria"], r["estoque_atual"],
                     r["valor_total_venda"], r["preco_venda_atual"], r["custo"]])
    _formatar(aba2)

    wb.save(SAIDA)
    return SAIDA, len(pesquisar), len(propria)


def main() -> None:
    conn = db.connect()
    rodada_id = conn.execute("SELECT MAX(id) FROM rodada").fetchone()[0]
    caminho, n_pesquisar, n_propria = exportar(conn, rodada_id)
    print(f"rodada {rodada_id}: {n_pesquisar} para pesquisar + {n_propria} de marca própria -> {caminho}")


if __name__ == "__main__":
    main()
