from collections import Counter, defaultdict
from pathlib import Path
import re
import unicodedata

from openpyxl import load_workbook


INPUT = Path(r"C:\Users\docze\Downloads\PERFUMARIA CLASSIFICADO.xlsx")
OUTPUT = Path("PERFUMARIA CLASSIFICADO - COM CATEGORIAS.xlsx")


def normalize(value):
    text = "" if value is None else str(value)
    text = unicodedata.normalize("NFD", text).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", text.lower()).strip()


def classify(text):
    if re.search(r"\bfralda|\bpants\b", text):
        return "PERFUMARIA > FRALDAS"
    if re.search(r"plenitud|roupa intima", text):
        return "PERFUMARIA > FRALDAS"
    if re.search(r"tintura|colorac|tonaliz|henna|agua oxig|oxidante|revelador|cor e ton", text):
        return "PERFUMARIA > TINTURAS"
    if re.search(
        r"maquiagem|\bbatom\b|\bgloss\b|\besmalte\b|\besm\b|\bbase\b|corretivo|po (?:compacto|solto)|\bblush\b|rimel|mascara (?:para )?cilios|delineador|\bdelin\b|\bsombra\b|lapis (?:de )?(?:olhos|labial|boca)|iluminador|primer|fixador|unha|cilios posticos|lip oil|sobrancelha|paleta|blindagem|bruma|skin match|shake scrub|\bbauny\b|by bandeja",
        text,
    ):
        return "PERFUMARIA > MAQUIAGEM"
    if re.search(
        r"creme dental|pasta dental|\bdental\b|\bdent\b|escova (?:de )?dente|\besc dent\b|fio dental|enxaguante bucal|bochecho|higiene bucal|oral[- ]?b|colgate|listerine|protes[ea] dent",
        text,
    ):
        return "PERFUMARIA > HIGIENE BUCAL"
    if re.search(r"\bbebe\b|\bbaby\b|infantil|crianca|\bkids\b|johnson.?s baby|joao e maria|hipoglos|(?:toalha|lenco) um[ei]d.*(?:rn|recem nascido)", text) and ("giovanna baby" not in text or "kids" in text):
        return "PERFUMARIA > LINHA INFANTIL"
    if re.search(
        r"shampoo|\bsham\b|condicion|\bcond\b|capilar|cabel[oa]|mascara (?:capilar|de tratamento)|\bmasc\b|creme (?:de )?pentear|crm (?:de )?pent|leave[ -]?in|finalizador|hair|pomada modeladora|gel fixador|pente|pantene|siage",
        text,
    ):
        return "PERFUMARIA > LINHA CAPILAR"
    if re.search(
        r"desodorante|\bdes aero\b|\bds aero\b|colonia|\bperfume\b|fragrancia|body splash|antitranspirante|roll[- ]?on|aerosol",
        text,
    ):
        return "PERFUMARIA > COLONIAS & DESODORANTES"
    if re.search(
        r"\bderma\b|acne|facial|serum|anti ?idade|retinol|vitamina c|hialuron|niacinamida|skincare|agua micelar|esfoliante facial|bepantriz|cerave|la roche|vichy|neutrogena|enxofre|principia|eucerin|argila|clareador|\bureia\b|aquaphor",
        text,
    ):
        return "PERFUMARIA > DERMOCOSMETICOS"
    if re.search(
        r"protetor solar|prot(?:etor)? sol|\bfps\b|filtro solar|bronzeador|acelerador bronzeado|pos sol|hidrat|\bhidra\b|\bhidr\b|\bloc hid\b|creme corporal|locao corporal|oleo corporal|creme (?:para )?maos",
        text,
    ):
        return "PERFUMARIA > BRONZ & HIDRATANTES"
    if re.search(
        r"sabonete|\bsab\b|higiene intima|intimo|absorvente|algodao|cotonete|hastes|papel higienico|barbeador|lamina|aparelho de barbear|gillette|venus|depilat|antisseptico|alcool|agua boric|pedra hume|gel lubrif|k-?med|talco|talqueira|lenco um[ei]d|toalha um[ei]d|repelente|xo inseto|\bpreserv\b|barba|espuma barbear|cera depi",
        text,
    ):
        return "PERFUMARIA > HIGIENE PESSOAL"
    return "PERFUMARIA > HIGIENE PESSOAL"


workbook = load_workbook(INPUT)
worksheet = workbook["Planilha1"]
worksheet.cell(1, 13).value = "Classificação"

counts = Counter()
examples = defaultdict(list)
for row in range(2, worksheet.max_row + 1):
    names = [worksheet.cell(row, column).value for column in (2, *range(4, 13))]
    text = normalize(" ".join(str(name) for name in names if name not in (None, 0, "0")))
    category = classify(text)
    worksheet.cell(row, 13).value = category
    counts[category] += 1
    if len(examples[category]) < 12:
        examples[category].append(worksheet.cell(row, 2).value)

workbook.save(OUTPUT)
print(f"Arquivo: {OUTPUT.resolve()}")
for category, count in sorted(counts.items()):
    print(f"{category}: {count}")
    for example in examples[category]:
        print(f"  - {example}")
