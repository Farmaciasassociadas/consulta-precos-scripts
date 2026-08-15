"""Precifica a "Planilha de itens em estoque tratado" (relatorio do ERP) com o
motor do app e devolve um Excel com: estoque precificado, comparacao contra o
preco praticado e a lista dos EANs que nunca foram pesquisados.

Regras de leitura do relatorio do ERP (confirmadas contra os precos que a loja
mantem a mao em eans.txt, 171 itens em comum):
  Margem            = Preco de Venda / Preco Comp.Unit - 1  (a base unitaria do
                      ERP e' `Preco Comp.Unit` = Preco de Compra / Unidades por Cx.)
  Preco de Venda    = preco de tabela; nos medicamentos com PMC e' o proprio PMC
  Preco 2 / Preco 3 = NAO sao preco de venda. Em 923 dos 1.224 itens preenchidos
                      sao identicos ao custo unitario, e nenhum dos 171 itens
                      cruzados bate com o preco real da loja. Ficam so como
                      referencia na saida.
  Preco de Promocao = o preco realmente praticado. Dos 145 itens com promocao que
                      a loja tambem mantem em eans.txt, 73 batem exatamente e
                      nenhum bate com o Preco de Venda.
  Ult. Prc. Entrada = custo real da ultima NF (ja com ICMS-ST); vem na unidade da
                      nota, entao as vezes e' o preco da CAIXA -- tratado aqui

Uso:
    python estoque_tratado_precificar.py [--entrada X.xlsx] [--saida Y.xlsx]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path
from statistics import median

import pandas as pd

from estoque_tratado_classificar import (_conteudo_declarado, auditar_categoria,
                                         conteudo_coletado, resolver_unidade)

APP = Path(r"C:\Users\docze\ConsultaPrecosEAN")
sys.path.insert(0, str(APP))

DIR_DADOS = Path(__file__).resolve().parent / "dados"
# Um arquivo por levantamento, lidos em ordem de nome: o mais novo sobrescreve o
# EAN do mais antigo. Assim uma reconferencia (o Huggies M caiu de R$ 42,99 para
# R$ 32,80 entre 14 e 15/08) entra sem reescrever o historico da pesquisa.
ARQUIVOS_PESQUISA_WEB = sorted(DIR_DADOS.glob("pesquisa_web_*.csv"))

ARQUIVO_NOTAS = Path(
    r"G:\.shortcut-targets-by-id\1q0IRmUp06SR55V7qNb7wVLwWjEauQntR\DROGARIA\todas nfs.xls")

ENTRADA_PADRAO = Path(
    r"G:\.shortcut-targets-by-id\1q0IRmUp06SR55V7qNb7wVLwWjEauQntR\DROGARIA"
    r"\Planilha de itens em estoque tratado.xlsx"
)

COLUNAS = ["ean", "descricao", "estoque", "custo_unit_erp", "grupo", "preco_compra",
           "un_cx", "margem1", "preco1", "margem2", "preco2", "margem3", "preco3",
           "margem_prom", "preco_prom", "ult_entrada", "classe_terapeutica"]

# ICMS-ST no PR nao chega a dobrar o custo; acima disso a "ultima entrada" e'
# preco de caixa (ou erro de digitacao), nao custo unitario.
RAZAO_ST_MAX = 2.0


def norm(valor) -> str:
    digitos = "".join(c for c in str(valor or "") if c.isdigit())
    return digitos.lstrip("0") or digitos


# ---------------------------------------------------------------- classificacao

def mapear_grupo(grupo: str, permitidas: set[str]) -> str | None:
    """"GENERICO - RX" -> "GENERICO > RX" quando a categoria existe no app."""
    candidato = str(grupo or "").strip().replace(" - ", " > ")
    return candidato if candidato in permitidas else None


def classificar(df: pd.DataFrame, mestre: dict[str, str], permitidas: set[str]):
    """Categoria final por EAN + motivo. O grupo do ERP manda quando e' especifico
    (ele carrega o eixo GENERICO/SIMILAR/ETICO, que define o lucro-alvo); os
    grupos guarda-chuva do ERP ("PERFUMARIA - GERAL", "NAO IDENTIFICADO") nao
    mapeiam para nada e cedem lugar a classificacao da base mestre, que e' mais
    granular."""
    final, motivo, divergencia = [], [], []
    for _, r in df.iterrows():
        do_erp = mapear_grupo(r["grupo"], permitidas)
        do_mestre = mestre.get(r["k"])
        if do_erp:
            final.append(do_erp)
            motivo.append("grupo do ERP")
        elif do_mestre:
            final.append(do_mestre)
            motivo.append("base mestre do app (grupo do ERP e generico)")
        else:
            final.append("")
            motivo.append("sem classificacao no ERP nem na base mestre")
        divergencia.append(
            f"ERP={do_erp} / mestre={do_mestre}"
            if do_erp and do_mestre and do_erp != do_mestre else ""
        )
    return final, motivo, divergencia


# ----------------------------------------------------------------------- custo

def custo_real(r, custo_nf: dict[str, float]) -> tuple[float | None, str]:
    """Custo unitario com ST. Devolve (custo, observacao).

    A ultima entrada e' o custo bom (traz o ICMS-ST, que o `Preco Comp.Unit` do
    ERP nao traz), mas vem na unidade da NF: em caixa/display e nos fracionados
    ela as vezes esta em outra base que o preco de venda. Quando as duas fontes
    divergem por mais de RAZAO_ST_MAX, o custo das NF-e ja consolidadas pelo app
    desempata; sem ele, fica o do ERP e o item sai sinalizado.
    """
    ult, unit, compra, un = r["ult_entrada"], r["custo_unit_erp"], r["preco_compra"], r["un_cx"]
    nf = custo_nf.get(r["k"])
    if not ult or ult <= 0:
        if unit and unit > 0:
            return unit, "sem ultima entrada: usado o custo unitario do ERP (pode nao incluir ICMS-ST)"
        return (nf, "custo veio das NF-e do app") if nf else (None, "sem custo")
    if unit and unit > 0 and not (1 / RAZAO_ST_MAX <= ult / unit <= RAZAO_ST_MAX):
        if un > 1 and abs(ult - compra) <= 0.02 * max(compra, 0.01):
            return ult / un, (f"ultima entrada veio em caixa (R$ {ult:.2f} / {un} un) "
                              f"-- convertida para unidade")
        alerta = (f"ultima entrada (R$ {ult:.2f}) e o custo unitario do ERP (R$ {unit:.2f}) "
                  f"diferem {max(ult, unit) / min(ult, unit):.0f}x -- bases de embalagem diferentes")
        if nf:
            escolha = ult if abs(ult - nf) <= abs(unit - nf) else unit
            return escolha, f"{alerta}; NF-e do app (R$ {nf:.2f}) desempatou em R$ {escolha:.2f}"
        return unit, f"{alerta}; sem NF-e para desempatar, mantido o do ERP -- CONFERIR"
    return ult, ""


# ------------------------------------------------------------------- severidade

def classificar_desvio(desvio_pct: float | None, impacto: float | None) -> str:
    if desvio_pct is None:
        return "SEM SUGESTAO"
    a = abs(desvio_pct)
    if a >= 0.30:
        return "ANALISE URGENTE"
    if a >= 0.15:
        return "REVISAR"
    if a >= 0.07:
        return "ATENCAO"
    return "PRECO ACEITAVEL"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--entrada", default=str(ENTRADA_PADRAO))
    ap.add_argument("--saida", default=None)
    args = ap.parse_args()

    entrada = Path(args.entrada)
    saida = Path(args.saida) if args.saida else entrada.with_name(
        entrada.stem + " - PRECIFICADO.xlsx")

    # "Barras" vem como texto no relatorio e algumas linhas tem zero a esquerda
    # (GTIN-8/UPC preenchido a 13). Sem forcar str o pandas infere int64 e come
    # o zero, e o codigo colado de volta no ERP nao acha o produto.
    df = pd.read_excel(entrada, dtype={"Barras": str})
    df.columns = COLUNAS
    df["k"] = df["ean"].map(norm)
    df["origem_linha"] = "relatorio de estoque"
    df = acrescentar_itens_das_notas(df)
    df["origem_linha"] = df["origem_linha"].fillna("so nas notas (fora do estoque atual)")

    import calcular_preco_sugerido as motor
    from config_app import ARQUIVO_CATEGORIAS_DISPONIVEIS, ARQUIVO_CATEGORIAS_PRODUTOS
    from dados_compartilhados import carregar_categorias_produtos, carregar_opcoes_categoria

    permitidas = set(carregar_opcoes_categoria(ARQUIVO_CATEGORIAS_DISPONIVEIS))
    mestre = {norm(e): c for e, c in
              carregar_categorias_produtos(ARQUIVO_CATEGORIAS_PRODUTOS).items()
              if c in permitidas}

    df["categoria_final"], df["origem_categoria"], df["divergencia_classificacao"] = \
        classificar(df, mestre, permitidas)

    # Auditoria: o grupo do ERP pode contradizer a classe terapeutica do proprio
    # ERP, o segmento do Brick ou o PMC da CMED. Quando contradiz, a evidencia
    # externa manda -- o eixo escolhe o lucro-alvo, entao errar aqui custa margem.
    referencia = motor.carregar_referencia()
    df["segmento_brick"] = df["k"].map(lambda k: (referencia.get(k) or {}).get("segmento"))
    df["pmc_cmed"] = df["k"].map(lambda k: (referencia.get(k) or {}).get("pmc"))
    auditoria = df.apply(auditar_categoria, axis=1, permitidas=permitidas)
    df["categoria_anterior"] = df["categoria_final"]
    df["categoria_proposta"] = [c for c, _ in auditoria]
    df["evidencia_categoria"] = [e for _, e in auditoria]
    df["categoria_final"] = df["categoria_proposta"].fillna(df["categoria_anterior"])
    print(f"Categorias adequadas: {df['categoria_proposta'].notna().sum()} "
          f"(restam {(df['categoria_final'].fillna('') == '').sum()} sem categoria)")

    custo_nf = motor.carregar_custos()
    custos = df.apply(custo_real, axis=1, custo_nf=custo_nf)
    df["custo_real"] = [c for c, _ in custos]
    df["obs_custo"] = [o for _, o in custos]

    # Preco praticado: a promocao e' o preco de balcao; o Preco de Venda e' tabela/PMC.
    df["preco_praticado"] = df["preco_prom"].where(df["preco_prom"] > 0, df["preco1"])
    df["base_preco_praticado"] = df["preco_prom"].gt(0).map(
        {True: "Preco de Promocao", False: "Preco de Venda"})

    # A nota fiscal manda no custo -- roda antes do motor para ele precificar em
    # cima do custo validado, nao do campo do cadastro.
    df = validar_custo_com_notas(df)

    # ---- motor do app, com a classificacao e o custo desta planilha
    custo_planilha = {r["k"]: float(r["custo_real"]) for _, r in df.iterrows()
                      if r["custo_real"] and r["custo_real"] > 0}
    motor.carregar_custos = lambda: {**custo_nf, **custo_planilha}
    meus = {r["k"]: float(r["preco_praticado"]) for _, r in df.iterrows()
            if r["preco_praticado"] and r["preco_praticado"] > 0}
    motor.carregar_meus_precos = lambda: meus

    override = {r["k"]: (r["categoria_final"] or None) for _, r in df.iterrows()}
    resultados, contagem = motor.calcular_todos(
        categorias_override=override, somente_eans=set(df["k"]))
    print(f"EANs precificados: {len(resultados)}")
    for status, n in sorted(contagem.items(), key=lambda x: -x[1]):
        print(f"  {status:45s} {n:5d}")

    saidas = pd.DataFrame(resultados).rename(columns={
        "ean": "k", "custo": "custo_usado_motor", "status": "status_motor",
        "tier": "tier", "piso": "piso", "alvo": "alvo",
        "preco_sugerido": "preco_sugerido", "justificativa": "justificativa"})
    df = df.merge(saidas, on="k", how="left")

    # ---- quem nunca foi pesquisado
    obs = motor.carregar_observacoes_por_ean()
    from precificacao.engine import mercado as eng_mercado
    tentativas = {e: len(v) for e, v in obs.items()}
    validos = {e: sum(1 for o in v if o.status in eng_mercado.STATUS_PRECO_VALIDO
                      and o.preco and o.preco > 0) for e, v in obs.items()}
    df["coletas"] = df["k"].map(tentativas).fillna(0).astype(int)
    df["concorrentes_com_preco"] = df["k"].map(validos).fillna(0).astype(int)
    marca_propria = motor.carregar_marca_propria()
    df["marca_propria"] = df["k"].isin(marca_propria).map({True: "SIM", False: ""})

    # Faixa crua do mercado coletado, por loja. A mediana sozinha mente em item
    # de marca com promocao agressiva: Aradois 50mg tem PMC ~R$ 57 e aparece a
    # R$ 5 nas grandes redes, entao a mediana da coleta (R$ 8,99) acusava a loja
    # de estar 5x cara praticando o proprio preco de tabela. A faixa mostra a
    # dispersao real e nao inventa um centro que loja nenhuma pratica.
    faixa = {e: sorted(median(v) for v in por_loja.values())
             for e, v in obs.items()
             if (por_loja := agrupar_por_loja(v, eng_mercado.STATUS_PRECO_VALIDO))}
    df["mercado_min"] = df["k"].map(lambda k: faixa[k][0] if k in faixa else None)
    df["mercado_max"] = df["k"].map(lambda k: faixa[k][-1] if k in faixa else None)
    df["lojas_na_faixa"] = df["k"].map(lambda k: len(faixa.get(k, ())))
    df["dentro_da_faixa_coletada"] = (
        df["lojas_na_faixa"].ge(2)
        & df["preco_praticado"].between(df["mercado_min"], df["mercado_max"]))

    df = conferir_apresentacao(df)

    # ---- base de embalagem (preco de caixa lancado como preco de unidade)
    unidades = df.apply(resolver_unidade, axis=1)
    df["fator_embalagem"] = [f for f, _, _, _ in unidades]
    df["preco_unitario_corrigido"] = [p for _, p, _, _ in unidades]
    df["situacao_unidade"] = [s for _, _, s, _ in unidades]
    df["explicacao_unidade"] = [e for _, _, _, e in unidades]
    resolvidos = df["situacao_unidade"].eq("RESOLVIDO")
    coleta_caixa = df["situacao_unidade"].eq("SUGESTAO EM BASE DE CAIXA")
    print(f"Base de embalagem: {resolvidos.sum()} preco da loja corrigido, "
          f"{coleta_caixa.sum()} sugestao trazida para a unidade, "
          f"{df['situacao_unidade'].eq('CONFERIR NA LOJA').sum()} para conferir, "
          f"{df['situacao_unidade'].eq('SEM PRECO').sum()} sem preco no ERP")
    # Corrigido o preco, a comparacao passa a ser contra o preco unitario real.
    df.loc[resolvidos, "preco_praticado"] = df.loc[resolvidos, "preco_unitario_corrigido"]
    df.loc[resolvidos, "base_preco_praticado"] = "corrigido para base unitaria"
    df.loc[coleta_caixa, "preco_sugerido"] = (
        df.loc[coleta_caixa, "preco_sugerido"] / df.loc[coleta_caixa, "fator_embalagem"]).round(2)
    df.loc[coleta_caixa, "justificativa"] = (
        "Sugestao dividida pelo fator de embalagem -- a coleta pegou a caixa. "
        + df.loc[coleta_caixa, "justificativa"].fillna(""))

    df = ancorar_pesquisa_web(df)

    # ---- comparacao
    df["desvio_pct"] = (df["preco_praticado"] - df["preco_sugerido"]) / df["preco_sugerido"]
    df.loc[df["preco_sugerido"].isna() | (df["preco_sugerido"] <= 0), "desvio_pct"] = pd.NA
    df["dif_unit"] = df["preco_praticado"] - df["preco_sugerido"]
    df["impacto_estoque"] = df["dif_unit"] * df["estoque"]
    df["valor_estoque"] = df["custo_real"] * df["estoque"]
    df["margem_atual_pct"] = (df["preco_praticado"] / df["custo_real"] - 1).where(df["custo_real"] > 0)
    df["margem_sugerida_pct"] = (df["preco_sugerido"] / df["custo_real"] - 1).where(df["custo_real"] > 0)
    df["direcao"] = df["desvio_pct"].map(
        lambda v: "" if pd.isna(v) else ("ACIMA do mercado" if v > 0 else "ABAIXO do mercado"))
    df["diagnostico"] = df.apply(diagnosticar, axis=1)
    df["classificacao_preco"] = [classificar_desvio(d, i) for d, i in
                                 zip(df["desvio_pct"], df["impacto_estoque"])]
    # Abaixo do custo e' sempre urgente, por menor que seja o desvio contra a
    # sugestao -- exceto onde a pesquisa mostrou que o custo cadastrado e' que
    # esta errado (custo com preco de varejo, custo em base de display). Ali o
    # problema e' o cadastro, nao o preco de venda.
    df.loc[df["custo_real"].gt(0) & df["preco_praticado"].le(df["custo_real"])
           & df["custo_confiavel"], "classificacao_preco"] = "ANALISE URGENTE"
    # Preco dentro da faixa observada no mercado e' preco bom, venha de onde
    # vier a sugestao: e' observacao direta, nao estimativa.
    df.loc[df["dentro_da_faixa_pesquisada"], "classificacao_preco"] = "PRECO ACEITAVEL"
    # Idem para a faixa da propria coleta: entre o menor e o maior concorrente o
    # preco esta no mercado, mesmo longe da mediana. E' o caso dos itens de marca
    # com promocao agressiva (Aradois, Allexofedrin, Clindamin-C), que enchiam a
    # fila de urgencia praticando preco de tabela legitimo.
    df.loc[df["dentro_da_faixa_coletada"], "classificacao_preco"] = "PRECO ACEITAVEL"
    # Apresentacao divergente invalida a comparacao inteira: nao da para dizer se
    # o preco esta certo comparando avulso com caixa. Vai para uma fila propria,
    # de conferencia de cadastro, em vez de virar acao de preco (Minilax C/1
    # contra a caixa de 7, que aparecia como o maior "prejuizo" da planilha).
    df.loc[df["divergencia_apresentacao"].ne(""),
           "classificacao_preco"] = "CONFERIR APRESENTACAO"
    # Item sem preco de venda nao esta mal precificado -- esta sem cadastro. Como
    # o desvio contra a sugestao da -100%, ele entrava em ANALISE URGENTE e
    # enchia a fila de acao com trabalho de cadastro, nao de preco.
    df.loc[df["preco_praticado"].fillna(0) <= 0, "classificacao_preco"] = "SEM PRECO CADASTRADO"

    # Dinheiro em jogo no estoque de hoje, nos dois sentidos: margem que esta
    # sendo entregue de graca (negativo) e venda que esta sendo perdida por preco
    # alto (positivo). E' por este valor que a aba de comparacao e' ordenada.
    df["prioridade"] = df["impacto_estoque"].abs()

    gravar(df, saida)
    print(f"Gravado: {saida}")


def diagnosticar(r) -> str:
    """Explica os desvios grandes -- quase sempre unidade ou custo, nao politica."""
    notas = []
    if r["un_cx"] > 1:
        notas.append(f"produto de caixa/display ({r['un_cx']} un): confira se o preco "
                     f"coletado na web e da unidade ou da caixa")
    if r["obs_custo"]:
        notas.append(r["obs_custo"])
    if r["concorrentes_com_preco"] == 0:
        notas.append("sem nenhum concorrente coletado: sugestao vem de custo/Brick")
    elif r["concorrentes_com_preco"] < 3:
        notas.append(f"so {r['concorrentes_com_preco']} concorrente(s) coletado(s)")
    if r["preco_prom"] > 0 and r["preco1"] > 0 and r["preco_prom"] > r["preco1"]:
        notas.append("promocao acima do preco de tabela: revisar cadastro")
    if r["custo_real"] and r["custo_real"] > 0 and r["preco_praticado"] <= r["custo_real"]:
        notas.append("PRECO ABAIXO DO CUSTO")
    if r["divergencia_classificacao"]:
        notas.append(f"classificacao divergente ({r['divergencia_classificacao']})")
    if r["divergencia_apresentacao"]:
        notas.insert(0, r["divergencia_apresentacao"])
    elif r["dentro_da_faixa_coletada"]:
        notas.insert(0, f"preco DENTRO da faixa de {r['lojas_na_faixa']:.0f} concorrentes "
                        f"(R$ {r['mercado_min']:.2f}-{r['mercado_max']:.2f})")
    # Desvio de varias vezes nao e' politica de preco errada: e' preco de caixa
    # contra preco de unidade em algum dos dois lados.
    p, s = r["preco_praticado"], r["preco_sugerido"]
    if p and s and s > 0 and not (0.2 < p / s < 5):
        notas.insert(0, f"PROVAVEL ERRO DE UNIDADE: praticado e sugerido diferem "
                        f"{max(p, s) / min(p, s):.0f}x (caixa x unidade/fracao)")
    return " | ".join(notas)


COLS_PRINCIPAL = ["ean", "descricao", "origem_linha", "grupo", "categoria_final", "origem_categoria",
                  "divergencia_classificacao", "classe_terapeutica", "estoque", "un_cx",
                  "custo_unit_erp", "ult_entrada", "custo_real", "custo_validado",
                  "custo_nf", "data_ultima_nf", "obs_custo",
                  "preco1", "preco_prom", "preco2", "preco_praticado", "base_preco_praticado",
                  "preco_sugerido", "origem_sugestao", "preco_sugerido_motor",
                  "piso", "alvo", "tier", "status_motor",
                  "concorrentes_com_preco", "justificativa"]

COLS_COMPARACAO = ["ranking", "classificacao_preco", "ean", "descricao", "categoria_final", "estoque",
                   "custo_real", "preco_praticado", "base_preco_praticado", "preco_sugerido",
                   "origem_sugestao", "mercado_min", "mercado_max", "lojas_na_faixa",
                   "dif_unit", "desvio_pct", "impacto_estoque", "margem_atual_pct",
                   "margem_sugerida_pct", "direcao", "diagnostico", "status_motor",
                   "concorrentes_com_preco", "justificativa"]

COLS_APRESENTACAO = ["ean", "descricao", "un_cx", "conteudo_erp", "conteudo_coletado",
                     "preco_praticado", "preco_sugerido", "mercado_min", "mercado_max",
                     "lojas_na_faixa", "estoque", "divergencia_apresentacao"]

ORDEM = ["ANALISE URGENTE", "REVISAR", "ATENCAO", "CONFERIR APRESENTACAO",
         "PRECO ACEITAVEL", "SEM PRECO CADASTRADO", "SEM SUGESTAO"]


def acrescentar_itens_das_notas(df: pd.DataFrame) -> pd.DataFrame:
    """Junta os EANs que a loja COMPROU mas que nao estao no relatorio de estoque.

    O relatorio de estoque e' uma foto: item que zerou, que chegou depois do
    fechamento ou que ainda nao foi cadastrado nao aparece nele -- e some da
    analise mesmo tendo nota de compra. Entram com estoque 0 e sem preco de
    venda, que e' exatamente a informacao util: comprei, e nao sei por quanto
    vendo.
    """
    if not ARQUIVO_NOTAS.exists():
        return df
    from custo_das_notas import custo_por_ean, ler_entradas

    notas = custo_por_ean(ler_entradas(ARQUIVO_NOTAS), norm)
    faltantes = [(k, v) for k, v in notas.items() if k not in set(df["k"])]
    if not faltantes:
        return df
    novos = pd.DataFrame([{
        "ean": k, "descricao": v["descricao_nf"], "estoque": 0,
        "custo_unit_erp": v["custo_nf"], "grupo": "", "preco_compra": v["custo_nf"],
        "un_cx": 1, "margem1": 0.0, "preco1": 0.0, "margem2": 0.0, "preco2": 0.0,
        "margem3": 0.0, "preco3": 0.0, "margem_prom": 0.0, "preco_prom": 0.0,
        "ult_entrada": v["custo_nf"], "classe_terapeutica": None, "k": k,
    } for k, v in faltantes])
    print(f"Itens so nas notas (comprados, fora do relatorio de estoque): {len(novos)}")
    return pd.concat([df, novos], ignore_index=True)


def validar_custo_com_notas(df: pd.DataFrame) -> pd.DataFrame:
    """Confere o custo de cada item contra o relatorio de entradas por nota.

    A NF e' o documento fiscal: o custo dela (total do item / quantidade, ja com
    ICMS-ST, IPI e frete) manda sobre qualquer campo do cadastro. Duas ressalvas:

    - Quando a NF vem em caixa e a loja vende a unidade, o custo da nota e' N
      vezes o unitario. Reconhecido pelo fator inteiro; ali o custo do cadastro
      e' que esta na base certa e a NF so confirma o total.
    - Sem NF no periodo nao ha o que validar. Nao e' sinal de custo bom: e'
      exatamente onde moram os custos que a pesquisa de mercado desmentiu.
    """
    if not ARQUIVO_NOTAS.exists():
        df["custo_validado"] = "SEM RELATORIO DE NOTAS"
        return df
    from custo_das_notas import custo_por_ean, ler_entradas

    notas = custo_por_ean(ler_entradas(ARQUIVO_NOTAS), norm)
    for campo in ("custo_nf", "unitario_bruto_nf", "tem_icms_st", "data_ultima_nf",
                  "fornecedor", "qtde_comprada", "n_notas"):
        df[campo] = df["k"].map(lambda k, c=campo: (notas.get(k) or {}).get(c))

    status, adotado, obs = [], [], []
    for _, r in df.iterrows():
        custo, nf = r["custo_real"], r["custo_nf"]
        if nf is None or pd.isna(nf):
            # Sem nota nao ha como validar, mas ha como reprovar: custo acima do
            # proprio preco de balcao nao e' margem negativa, e' cadastro podre
            # (Vick Vaporub 12g com R$ 440 de custo). Sem esta marca eles entram
            # no motor e viram sugestao absurda.
            praticado = r["preco_praticado"]
            if custo and praticado and praticado > 0 and custo > praticado * 1.05:
                status.append("SEM NF - CUSTO IMPOSSIVEL")
                obs.append(f"custo R$ {custo:.2f} acima do proprio preco de venda "
                           f"R$ {praticado:.2f} e sem NF que comprove: descartado")
                # Descartar de verdade: mantido, ele vira piso de margem e empurra
                # a sugestao para cima ignorando a concorrencia -- a Nimesulida
                # C/12 saia a R$ 25,99 com 5 concorrentes entre R$ 8 e R$ 14.
                # Sem custo, o motor precifica so pelo mercado, que e' o certo
                # quando o custo e' o dado ruim.
                adotado.append(None)
            else:
                status.append("SEM NF NO PERIODO")
                obs.append("custo nao validado: nenhuma compra deste EAN no relatorio de notas")
                adotado.append(custo)
            continue
        if not custo or custo <= 0:
            status.append("ADOTADO DA NF")
            adotado.append(nf)
            obs.append(f"cadastro sem custo; adotado o da NF de {r['data_ultima_nf']}")
            continue
        razao = nf / custo
        inteiro = round(razao)
        if 0.98 <= razao <= 1.02:
            status.append("CONFIRMADO PELA NF")
            adotado.append(nf)
            obs.append("")
        elif inteiro >= 2 and abs(razao - inteiro) <= 0.05 * inteiro:
            status.append("NF EM BASE DE CAIXA")
            adotado.append(custo)
            obs.append(f"a nota traz a embalagem de {inteiro} un (R$ {nf:.2f}); "
                       f"o custo unitario R$ {custo:.2f} esta correto")
        elif razao > 1.5:
            # Diferenca desta ordem sem fator inteiro limpo e' quase sempre
            # embalagem (fardo de 6 com desconto nao da razao 6,00 exata).
            # Adotar a nota aqui poria um sabonete com R$ 22 de custo.
            status.append("CONFERIR EMBALAGEM")
            adotado.append(custo)
            obs.append(f"NF R$ {nf:.2f} e' {razao:.1f}x o cadastro R$ {custo:.2f} sem fator "
                       f"inteiro: provavel compra em fardo. Mantido o custo do cadastro")
        else:
            status.append("DIVERGENTE DA NF")
            adotado.append(nf)
            obs.append(f"cadastro R$ {custo:.2f} contra NF R$ {nf:.2f} "
                       f"({razao:.2f}x): adotado o da nota")
    df["custo_validado"] = status
    df["custo_anterior"] = df["custo_real"]
    df["custo_real"] = adotado
    df["obs_validacao_custo"] = obs
    print("Validacao de custo contra as notas: "
          + " | ".join(f"{k} {v}" for k, v in df["custo_validado"].value_counts().items()))
    return df


# A guarda de 3x sobre a mediana do mercado mora em
# precificacao.engine.mercado.excede_sanidade, aplicada dentro do motor -- assim
# vale para o painel do app e para a rodada SQLite, nao so para esta planilha. A
# versao que existia aqui limitava a sugestao a 3x; o motor DESCARTA, porque
# Paracetamol limitado a 3 x R$ 5,99 continua sendo um preco errado que alguem
# aplica. O falso positivo do Aradois (PMC R$ 57 contra mediana promocional de
# R$ 8,99) e' resolvido la pela isencao de PMC, sem precisar excluir os EANs
# pesquisados a mao.


def _num(valor) -> float | None:
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return None
    return None if v != v or v <= 0 else v


def agrupar_por_loja(observacoes, status_validos) -> dict[str, list[float]]:
    """Precos validos por site. Uma loja coletada 40 vezes conta como uma loja."""
    por_loja: dict[str, list[float]] = {}
    for o in observacoes:
        if o.status in status_validos and o.preco and o.preco > 0:
            por_loja.setdefault(o.site, []).append(o.preco)
    return por_loja


def conferir_apresentacao(df: pd.DataFrame) -> pd.DataFrame:
    """Compara o conteudo da embalagem do ERP com o que a COLETA viu no site.

    Enquanto a coleta so grava (ean, preco), avulso e caixa entram na mesma
    chave e viram desvio de margem fantasma. O nome do produto ja vem junto no
    precos.csv e diz a apresentacao -- basta ler. Sai mais barato que mudar o
    schema da coleta e os 8 userscripts, e vale retroativamente para tudo que ja
    foi coletado.
    """
    nomes = pd.read_csv(APP / "precos.csv", encoding="utf-8-sig", low_memory=False,
                        usecols=["ean", "status", "nome"], dtype=str)
    nomes = nomes[nomes["status"].eq("OK") & nomes["nome"].notna()]
    por_ean = nomes.groupby(nomes["ean"].map(norm))["nome"].apply(list)
    df["conteudo_coletado"] = df["k"].map(por_ean.map(conteudo_coletado))
    df["conteudo_erp"] = df["descricao"].map(lambda d: _conteudo_declarado(d, minimo=1))
    a, b = df["conteudo_erp"], df["conteudo_coletado"]
    df["divergencia_apresentacao"] = [
        "" if pd.isna(x) or pd.isna(y) or x == y else
        f"ERP vende {x:.0f} un e a coleta leu {y:.0f} un no site "
        f"({max(x, y) / min(x, y):.0f}x): preco coletado nao e comparavel"
        for x, y in zip(a, b)]
    n = df["divergencia_apresentacao"].ne("").sum()
    print(f"Apresentacao: {b.notna().sum()} EANs com conteudo lido na coleta, "
          f"{n} divergem do conteudo do ERP")
    return df


def ler_pesquisa_web() -> pd.DataFrame:
    """Todos os levantamentos manuais, o mais novo mandando no EAN repetido."""
    partes = [pd.read_csv(a, sep=";", decimal=",", dtype={"ean": str}, encoding="utf-8-sig")
              for a in ARQUIVOS_PESQUISA_WEB]
    web = pd.concat(partes, ignore_index=True)
    web["k"] = web["ean"].map(norm)
    return web.drop_duplicates("k", keep="last")


def faixas_pesquisadas() -> dict[str, tuple[float, float]]:
    """{ean normalizado: (min, max)} dos levantamentos manuais."""
    return {norm(r["ean"]): (r["preco_mercado_min"], r["preco_mercado_max"])
            for _, r in ler_pesquisa_web().iterrows()}


def ancorar_pesquisa_web(df: pd.DataFrame) -> pd.DataFrame:
    """Traz a sugestao para dentro da faixa de preco levantada na internet.

    Nesses EANs o motor nao tinha como acertar: ou o custo cadastrado esta errado
    (e ele multiplicou o erro pelo markup), ou a coleta nao trouxe concorrente
    nenhum e ele caiu em custo x markup, ou o EAN curto casou com outro produto.
    A faixa pesquisada e' observacao direta do mercado -- vale mais que qualquer
    das tres. Quando a sugestao do motor ja cai dentro da faixa, ela fica: os
    dois concordam e nao ha o que ancorar.
    """
    web = ler_pesquisa_web()
    faixa = {norm(r["ean"]): (r["preco_mercado_min"], r["preco_mercado_max"])
             for _, r in web.iterrows()}
    # A propria pesquisa diz quando o custo do ERP nao vale: nesses itens a
    # comparacao "preco abaixo do custo" e' falsa, o custo e' que esta alto.
    custo_suspeito = {norm(r["ean"]) for _, r in web.iterrows()
                      if "CUSTO" in str(r["causa_raiz"]).upper()}
    df["preco_mercado_min"] = df["k"].map(lambda k: faixa.get(k, (None, None))[0])
    df["preco_mercado_max"] = df["k"].map(lambda k: faixa.get(k, (None, None))[1])
    df["custo_confiavel"] = ~df["k"].isin(custo_suspeito)
    df["dentro_da_faixa_pesquisada"] = (
        df["preco_mercado_min"].notna()
        & df["preco_praticado"].between(df["preco_mercado_min"], df["preco_mercado_max"]))
    df["preco_sugerido_motor"] = df["preco_sugerido"]
    df["origem_sugestao"] = "motor"
    ancorados = 0
    for i, r in df.iterrows():
        limites = faixa.get(r["k"])
        if not limites or pd.isna(r["preco_sugerido"]):
            continue
        minimo, maximo = limites
        alvo = min(max(float(r["preco_sugerido"]), minimo), maximo)
        if abs(alvo - r["preco_sugerido"]) <= 0.005:
            df.at[i, "origem_sugestao"] = "motor (confirmado pela pesquisa web)"
            continue
        df.at[i, "preco_sugerido"] = round(alvo, 2)
        df.at[i, "origem_sugestao"] = (
            f"faixa pesquisada na web (R$ {minimo:.2f}-{maximo:.2f}); "
            f"motor sugeria R$ {r['preco_sugerido']:.2f}")
        ancorados += 1
    print(f"Pesquisa web: {ancorados} sugestoes ancoradas na faixa de mercado, "
          f"{(df['origem_sugestao'] == 'motor (confirmado pela pesquisa web)').sum()} confirmadas, "
          f"{df['dentro_da_faixa_pesquisada'].sum()} com o preco da loja ja dentro da faixa")
    return df


def pesquisa_web(df: pd.DataFrame) -> pd.DataFrame:
    """Junta o preco de mercado levantado a mao na internet com o que a planilha diz.

    Os itens de maior desvio nao se resolvem com mais calculo: e' preciso saber o
    que o EAN realmente e' e por quanto o mercado vende. Estes CSVs sao esse
    levantamento -- ver ARQUIVOS_PESQUISA_WEB e a coluna `fonte` de cada linha.
    """
    web = ler_pesquisa_web()
    junta = web.merge(
        df[["k", "estoque", "custo_real", "preco_praticado", "preco_sugerido",
            "concorrentes_com_preco"]], on="k", how="left")
    junta["preco_mercado_medio"] = (junta["preco_mercado_min"] + junta["preco_mercado_max"]) / 2
    junta["erro_do_meu_preco"] = (junta["preco_praticado"] / junta["preco_mercado_medio"] - 1)
    junta["erro_da_sugestao"] = (junta["preco_sugerido"] / junta["preco_mercado_medio"] - 1)
    return junta[["ean", "descricao", "estoque", "custo_real", "preco_praticado",
                  "preco_sugerido", "preco_mercado_min", "preco_mercado_max",
                  "erro_do_meu_preco", "erro_da_sugestao", "concorrentes_com_preco",
                  "causa_raiz", "acao", "fonte"]]


COLUNAS_DINHEIRO = {"custo_unit_erp", "ult_entrada", "custo_real", "preco1", "preco_prom",
                    "preco2", "preco_praticado", "preco_sugerido", "preco_sugerido_motor",
                    "piso", "alvo", "dif_unit", "impacto_estoque", "impacto_total",
                    "preco_unitario_corrigido", "preco_mercado_min", "preco_mercado_max",
                    "mercado_min", "mercado_max", "pmc_cmed"}
COLUNAS_PCT = {"desvio_pct", "margem_atual_pct", "margem_sugerida_pct",
               "erro_do_meu_preco", "erro_da_sugestao", "desvio_medio"}
# Formato contabil do Excel: simbolo alinhado a esquerda, valor a direita, zero
# como traco. Sem isto o Excel mostra so o numero e o usuario nao ve a moeda.
FORMATO_CONTABIL = r'_-"R$" * #,##0.00_-;-"R$" * #,##0.00_-;_-"R$" * "-"??_-;_-@_-'
COR_CATEGORIA = {"ANALISE URGENTE": "FFC7CE", "REVISAR": "FFD9A0",
                 "ATENCAO": "FFEB9C", "PRECO ACEITAVEL": "C6EFCE",
                 "CONFERIR APRESENTACAO": "BDD7EE",
                 "SEM PRECO CADASTRADO": "D9D9D9"}


def formatar(caminho: Path) -> None:
    """Cabecalho fixo, filtro, largura e formato de numero -- a planilha e' para
    ler no Excel, e sem isto tudo sai como numero cru de 14 casas."""
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(caminho)
    cabecalho = Font(bold=True, color="FFFFFF")
    fundo = PatternFill("solid", fgColor="44546A")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        nomes = [c.value for c in ws[1]]
        for c in ws[1]:
            c.font, c.fill = cabecalho, fundo
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for i, nome in enumerate(nomes, start=1):
            letra = get_column_letter(i)
            largura = max(len(str(nome)) + 3, 10)
            if nome == "ean":
                # Texto, nao numero: com 13 digitos o Excel exibe 7,89601E+12 e
                # o copiar/colar leva a notacao cientifica junto.
                largura = 16
                for c in ws[letra][1:]:
                    c.number_format = "@"
                    c.alignment = Alignment(horizontal="left")
                    if c.value is not None:
                        c.value = str(c.value)
            elif nome in COLUNAS_DINHEIRO or nome in COLUNAS_PCT:
                formato = FORMATO_CONTABIL if nome in COLUNAS_DINHEIRO else "0.0%"
                largura = max(largura, 14)
                for c in ws[letra][1:]:
                    c.number_format = formato
            elif nome in ("descricao", "justificativa", "explicacao_unidade", "causa_raiz",
                          "acao", "diagnostico", "obs_custo", "evidencia_categoria", "fonte"):
                largura = 55 if nome == "descricao" else 70
            elif nome in ("categoria_final", "categoria_anterior", "categoria_proposta",
                          "grupo", "status_motor", "divergencia_classificacao"):
                largura = 32
            ws.column_dimensions[letra].width = min(largura, 70)
        if "classificacao_preco" in nomes:
            letra = get_column_letter(nomes.index("classificacao_preco") + 1)
            for c in ws[letra][1:]:
                cor = COR_CATEGORIA.get(c.value)
                if cor:
                    c.fill = PatternFill("solid", fgColor=cor)
    wb.save(caminho)


def gravar(df: pd.DataFrame, saida: Path) -> None:
    nunca = df[(df["coletas"] == 0) & (df["marca_propria"] != "SIM")]
    nunca_mp = df[(df["coletas"] == 0) & (df["marca_propria"] == "SIM")]
    mudou = df[df["categoria_proposta"].notna()]
    unidade = df[df["situacao_unidade"].ne("")]
    # Ordem unica: o maior dinheiro em jogo primeiro, independente da categoria.
    # Agrupar por categoria antes escondia o 3o maior impacto da planilha na
    # linha 483, atras de 480 itens urgentes de R$ 20. A severidade continua
    # visivel na coluna colorida e filtravel.
    comp = df[df["preco_sugerido"].notna()].copy()
    comp = comp.sort_values("prioridade", ascending=False)
    comp.insert(0, "ranking", range(1, len(comp) + 1))
    with pd.ExcelWriter(saida, engine="openpyxl") as xl:
        df[COLS_PRINCIPAL].to_excel(xl, sheet_name="Estoque precificado", index=False)
        comp[COLS_COMPARACAO].to_excel(xl, sheet_name="Comparacao meu preco", index=False)
        cols_nunca = ["ean", "descricao", "grupo", "categoria_final", "estoque",
                      "custo_real", "preco_praticado"]
        nunca[cols_nunca].to_excel(xl, sheet_name="Nunca pesquisados", index=False)
        nunca_mp[cols_nunca].to_excel(xl, sheet_name="Nunca pesq (marca propria)", index=False)
        mudou[["ean", "descricao", "grupo", "classe_terapeutica", "segmento_brick", "pmc_cmed",
               "categoria_anterior", "categoria_proposta", "evidencia_categoria",
               "estoque"]].to_excel(xl, sheet_name="Mudanca de categoria", index=False)
        unidade.sort_values("situacao_unidade")[
            ["situacao_unidade", "ean", "descricao", "un_cx", "custo_unit_erp", "ult_entrada",
             "custo_real", "preco1", "preco_prom", "fator_embalagem",
             "preco_unitario_corrigido", "preco_sugerido", "estoque", "explicacao_unidade"]
        ].to_excel(xl, sheet_name="Problemas de unidade", index=False)
        df[df["divergencia_apresentacao"].ne("")].sort_values(
            "valor_estoque", ascending=False)[COLS_APRESENTACAO].to_excel(
            xl, sheet_name="Apresentacao divergente", index=False)
        df[df["custo_validado"].ne("CONFIRMADO PELA NF")].sort_values(
            ["custo_validado", "valor_estoque"], ascending=[True, False])[
            ["custo_validado", "ean", "descricao", "estoque", "valor_estoque", "un_cx",
             "custo_unit_erp", "ult_entrada", "custo_anterior", "custo_nf", "custo_real",
             "data_ultima_nf", "fornecedor", "n_notas", "obs_validacao_custo"]
        ].to_excel(xl, sheet_name="Validacao de custo", index=False)
        pesquisa_web(df).to_excel(xl, sheet_name="Pesquisa web", index=False)
        df[df["divergencia_classificacao"].ne("")][
            ["ean", "descricao", "grupo", "categoria_final", "divergencia_classificacao",
             "classe_terapeutica", "estoque"]
        ].to_excel(xl, sheet_name="Divergencia com a mestre", index=False)
        resumo = (comp.groupby("classificacao_preco")
                  .agg(itens=("ean", "count"),
                       impacto_total=("impacto_estoque", "sum"),
                       desvio_medio=("desvio_pct", "mean"))
                  .reindex(ORDEM).dropna(how="all").reset_index())
        resumo.to_excel(xl, sheet_name="Resumo", index=False)
    formatar(saida)
    print(resumo.to_string(index=False))
    print(f"Nunca pesquisados: {len(nunca)}")


if __name__ == "__main__":
    main()
