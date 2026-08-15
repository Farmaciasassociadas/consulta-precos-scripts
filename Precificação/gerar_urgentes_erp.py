"""Lista unica e priorizada do que corrigir NO ERP, ordenada por dinheiro.

Nao e um relatorio de analise -- e a fila de digitacao. Cada linha diz o campo,
o valor de hoje, o valor a por e a evidencia. Item sem evidencia forte fica de
fora: fila cheia de "talvez" nao e usada.

Uso:
    python gerar_urgentes_erp.py [--planilha X.xlsx] [--saida Y.xlsx]
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

import pandas as pd

from gerar_acoes_cadastro import candidatos, dv_ok, digitos, norm

APP = Path(r"C:\Users\docze\ConsultaPrecosEAN")
sys.path.insert(0, str(APP))

PLANILHA_PADRAO = Path(
    r"G:\.shortcut-targets-by-id\1q0IRmUp06SR55V7qNb7wVLwWjEauQntR\DROGARIA"
    r"\Planilha de itens em estoque tratado - PRECIFICADO.xlsx")

DIR_DADOS = Path(__file__).resolve().parent / "dados"

CAMPO_PRECO = "Preco de Promocao (o preco praticado)"
CAMPO_CUSTO = "Ult. Prc. Entrada / Preco de Compra"

ARQUIVO_MARKUP = APP / "precificacao" / "dados" / "politica_markup.csv"
MARKUP_PADRAO = 56.2  # ETICOS/PADRAO, a linha 1 da politica


def carregar_web() -> dict[str, tuple[float, float, str, bool]]:
    """Faixas conferidas a mao na internet: (min, max, fonte, manter).

    Fonte unica com o `estoque_tratado_precificar.py`, que le os mesmos arquivos.
    Estava duplicado numa constante aqui dentro com 11 EANs enquanto os CSVs ja
    tinham 45 -- a fila deixava de fora 36 itens conferidos, e as duas listas
    iam divergir a cada rodada nova de pesquisa. Arquivo mais novo vence.

    `manter` marca o que a pesquisa classificou como FALSO POSITIVO: ali o preco
    de balcao esta certo e a faixa serve so de evidencia, nunca de correcao. Sem
    isto a Loratadina (praticado R$ 11,99, faixa ate R$ 11,07) entraria na fila
    mandando baixar R$ 0,92 num item que a propria pesquisa manda deixar quieto.
    """
    faixas: dict[str, tuple[float, float, str, bool]] = {}
    for arquivo in sorted(DIR_DADOS.glob("pesquisa_web_*.csv")):
        d = pd.read_csv(arquivo, sep=";", decimal=",", dtype={"ean": str},
                        encoding="utf-8-sig")
        for _, r in d.iterrows():
            if pd.isna(r["preco_mercado_min"]) or pd.isna(r["preco_mercado_max"]):
                continue
            manter = str(r.get("acao") or "").strip().upper().startswith("MANTER")
            faixas[norm(r["ean"])] = (float(r["preco_mercado_min"]),
                                      float(r["preco_mercado_max"]),
                                      str(r["fonte"])[:70], manter)
    return faixas


def carregar_markup() -> dict[str, float]:
    """markup_sobre_custo_pct por categoria, da mesma politica que o motor usa.

    O arquivo tem linhas com campo a mais (ponto-e-virgula dentro da observacao),
    que o parser em C rejeita; o engine='python' com on_bad_lines pula so essas.
    """
    d = pd.read_csv(ARQUIVO_MARKUP, sep=";", decimal=",", encoding="utf-8-sig",
                    engine="python", on_bad_lines="skip")
    return {str(r["categoria_exata"]).strip(): float(r["markup_sobre_custo_pct"])
            for _, r in d.iterrows() if pd.notna(r["markup_sobre_custo_pct"])}


def custo_estimado(r, markup: dict[str, float]) -> tuple[float | None, str]:
    """Custo a por no lugar do impossivel, e de onde ele saiu.

    Estimativa, nao leitura de nota -- por isso devolve junto a origem, para a
    linha da planilha dizer em que confiar. A ancora vem do melhor dado
    disponivel, nesta ordem: o menor concorrente coletado (preco real de
    mercado), a sugestao do motor (que nesses itens ja foi calculada SEM o custo
    podre, so pela concorrencia) e por fim o proprio preco de balcao.
    Dividida pelo markup da categoria, que e' a mesma politica que o motor
    aplica -- assim o custo estimado e o preco sugerido ficam coerentes.
    """
    pct = markup.get(str(r.get("categoria_final") or "").strip(), MARKUP_PADRAO)
    # Uma loja so nao e' mercado: a Atorvastatina 40 C/30 tinha um unico
    # concorrente a R$ 149,49 (contra sugestao de R$ 60,95 e balcao de R$ 37,49)
    # e o custo estimado saiu R$ 95,70. Com 2+ lojas o minimo ja e' um preco que
    # alguem de fato pratica; abaixo disso a sugestao do motor, que passou pelas
    # camadas de outlier, e' a ancora melhor.
    lojas = r.get("lojas_na_faixa")
    ancoras = [("preco_sugerido", "sugestao do motor (calculada so pelo mercado)"),
               ("preco_praticado", "preco de balcao de hoje")]
    if pd.notna(lojas) and lojas >= 2:
        ancoras.insert(0, ("mercado_min", f"menor de {lojas:.0f} concorrentes coletados"))
    for campo, nome in ancoras:
        v = r.get(campo)
        if pd.notna(v) and v and v > 0:
            return round(float(v) / (1 + pct / 100), 2), f"estimado de {nome} / markup {pct:.0f}%"
    return None, "sem ancora: confira a nota fiscal"


def carregar(planilha: Path) -> pd.DataFrame:
    df = pd.read_excel(planilha, sheet_name="Estoque precificado", dtype={"ean": str})
    comp = pd.read_excel(planilha, sheet_name="Comparacao meu preco", dtype={"ean": str})
    df = df.merge(comp[["ean", "mercado_min", "mercado_max", "lojas_na_faixa"]],
                  on="ean", how="left")
    apr = pd.read_excel(planilha, sheet_name="Apresentacao divergente", dtype={"ean": str})
    df["apresentacao_divergente"] = df["ean"].isin(apr["ean"])
    df["k"] = df["ean"].map(norm)
    return df


def acoes(df: pd.DataFrame, pmc: dict[str, float], markup: dict[str, float],
          web: dict[str, tuple[float, float, str, bool]]) -> pd.DataFrame:
    """Uma linha por acao. `dinheiro` e' o criterio de ordem: o que muda mais
    caixa vem primeiro, seja margem entregue de graca ou venda que nao acontece."""
    out = []

    def add(r, prio, acao, campo, de, para, motivo, evidencia, dinheiro):
        # preco de hoje, sugerido e faixa de mercado vao em TODA linha: quem
        # digita precisa ver o preco na mesma tela da correcao de custo ou de
        # EAN, senao tem de voltar na planilha grande para decidir.
        faixa_txt = (f"R$ {r['mercado_min']:.2f}-{r['mercado_max']:.2f} "
                     f"({r['lojas_na_faixa']:.0f} lojas)"
                     if pd.notna(r.get("mercado_min")) else "")
        out.append({
            "prioridade": prio, "acao": acao, "ean": r["ean"],
            "descricao": r["descricao"], "campo_no_erp": campo,
            "valor_hoje": round(float(de), 2) if de is not None else None,
            "valor_novo": round(float(para), 2) if para is not None else None,
            "preco_hoje": r["preco_praticado"], "preco_sugerido": r["preco_sugerido"],
            "faixa_de_mercado": faixa_txt,
            "estoque": r["estoque"], "dinheiro_em_jogo": round(float(dinheiro), 2),
            "motivo": motivo, "evidencia": evidencia})

    for _, r in df.iterrows():
        k, est = r["k"], float(r["estoque"] or 0)
        custo = r["custo_real"] if pd.notna(r["custo_real"]) else 0
        prat = r["preco_praticado"] if pd.notna(r["preco_praticado"]) else 0
        faixa = web.get(k)

        # 1. Custo impossivel: acima do proprio preco de venda e sem NF.
        #    O pipeline ja zera `custo_real` nesses (senao vira piso e empurra a
        #    sugestao), entao o valor que ainda esta NO ERP -- que e' o que a
        #    loja precisa corrigir -- so aparece em custo_unit_erp/ult_entrada.
        if r["custo_validado"] == "SEM NF - CUSTO IMPOSSIVEL":
            no_erp = max([v for v in (r["custo_unit_erp"], r["ult_entrada"])
                          if pd.notna(v) and v > 0], default=0)
            if no_erp > 0:
                novo, origem = custo_estimado(r, markup)
                fonte = (f"internet: R$ {faixa[0]:.2f}-{faixa[1]:.2f} em {faixa[2]}. "
                         if faixa else "")
                motivo = (f"custo R$ {no_erp:.2f} acima do proprio preco de venda "
                          f"(R$ {prat:.2f}) e sem NF que comprove")
                # Custo estimado a partir do mercado que ainda fica acima do
                # preco de balcao significa que o cadastro tem DOIS problemas:
                # o custo esta podre e o preco esta baixo demais. Sem este aviso
                # a linha manda por um custo maior que o preco e parece erro.
                if novo and prat > 0 and novo >= prat:
                    motivo += (f". ATENCAO: mesmo o custo estimado (R$ {novo:.2f}) fica "
                               f"acima do seu preco -- o preco tambem esta errado, "
                               f"veja a aba 2")
                add(r, 1, "CORRIGIR CUSTO", CAMPO_CUSTO, no_erp, novo, motivo,
                    f"{fonte}{origem} -- ESTIMATIVA, confira a nota fiscal",
                    no_erp * max(est, 1))

        # 2. Acima do PMC: teto legal, nao admite excecao.
        teto = pmc.get(k)
        if teto and prat > teto * 1.001:
            novo = min(r["preco_sugerido"], teto) if pd.notna(r["preco_sugerido"]) else teto
            add(r, 1, "BAIXAR PRECO (acima do PMC)", CAMPO_PRECO, prat, novo,
                f"venda {prat / teto - 1:.0%} acima do teto legal da CMED (PMC R$ {teto:.2f})",
                "tabela CMED", (prat - novo) * max(est, 1))

        # 3. Preco fora da faixa conferida na internet.
        if faixa and prat > 0 and not faixa[3]:
            minimo, maximo, fonte, _ = faixa
            if prat < minimo * 0.95:
                add(r, 2, "SUBIR PRECO", CAMPO_PRECO, prat, minimo,
                    f"vendendo {1 - prat / minimo:.0%} abaixo do menor concorrente",
                    f"R$ {minimo:.2f}-{maximo:.2f} em {fonte}", (minimo - prat) * est)
            elif prat > maximo * 1.05:
                add(r, 2, "BAIXAR PRECO", CAMPO_PRECO, prat, maximo,
                    f"vendendo {prat / maximo - 1:.0%} acima do maior concorrente",
                    f"R$ {minimo:.2f}-{maximo:.2f} em {fonte}", (prat - maximo) * est)

        # 4. Marca propria com custo em base de embalagem errada. NF e cadastro
        #    concordam, entao nenhuma guarda pega -- mas custo 2x+ o proprio
        #    preco de venda, repetido em toda a linha, e' fator de venda faltando.
        if (custo > 0 and prat > 0 and custo > prat * 1.8
                and r["custo_validado"] == "CONFIRMADO PELA NF"):
            add(r, 2, "CADASTRAR FATOR DE VENDA", "embalagens_produtos.csv (fator_venda)",
                custo, None,
                f"custo R$ {custo:.2f} e' {custo / prat:.1f}x o proprio preco de venda",
                "NF e cadastro concordam: a nota vem em outra base de embalagem",
                (custo - prat) * est)

        # 5. EAN que nao fecha o digito verificador.
        if not dv_ok(digitos(r["ean"])):
            novo = next((c for c, _ in candidatos(r["ean"]) if dv_ok(c)), None)
            if novo:
                add(r, 3, "CORRIGIR EAN", "Barras", None, None,
                    "codigo nao fecha o digito verificador: nunca casa com coleta nem CMED",
                    f"codigo provavel {novo}", custo * est)
                out[-1]["valor_hoje"] = r["ean"]
                out[-1]["valor_novo"] = novo

    a = pd.DataFrame(out)
    return a.sort_values(["prioridade", "dinheiro_em_jogo"], ascending=[True, False])


def formatar(caminho: Path) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    largura = {"prioridade": 11, "acao": 28, "ean": 16, "descricao": 46,
               "campo_no_erp": 34, "valor_hoje": 14, "valor_novo": 14,
               "preco_hoje": 13, "preco_sugerido": 15, "faixa_de_mercado": 24,
               "estoque": 10, "dinheiro_em_jogo": 17, "motivo": 62, "evidencia": 68}
    cor = {1: "FFC7CE", 2: "FFD9A0", 3: "FFEB9C"}
    wb = load_workbook(caminho)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        nomes = [c.value for c in ws[1]]
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="44546A")
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for i, nome in enumerate(nomes, start=1):
            letra = get_column_letter(i)
            ws.column_dimensions[letra].width = largura.get(nome, 16)
            if nome in ("ean", "valor_hoje", "valor_novo"):
                for c in ws[letra][1:]:
                    if isinstance(c.value, str):
                        c.number_format = "@"
                        c.alignment = Alignment(horizontal="left")
                    elif nome != "ean":
                        c.number_format = r'_-"R$" * #,##0.00_-;-"R$" * #,##0.00_-'
            elif nome in ("dinheiro_em_jogo", "preco_hoje", "preco_sugerido"):
                for c in ws[letra][1:]:
                    c.number_format = r'_-"R$" * #,##0.00_-;-"R$" * #,##0.00_-'
        if "prioridade" in nomes:
            letra = get_column_letter(nomes.index("prioridade") + 1)
            for c in ws[letra][1:]:
                if c.value in cor:
                    c.fill = PatternFill("solid", fgColor=cor[c.value])
    wb.save(caminho)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--planilha", default=str(PLANILHA_PADRAO))
    ap.add_argument("--saida", default=None)
    args = ap.parse_args()

    planilha = Path(args.planilha)
    saida = Path(args.saida) if args.saida else planilha.with_name("URGENTE - CORRIGIR NO ERP.xlsx")

    import calcular_preco_sugerido as motor
    pmc = {k: v["pmc"] for k, v in motor.carregar_referencia().items() if v.get("pmc")}

    df = carregar(planilha)
    web = carregar_web()
    print(f"Pesquisa web: {len(web)} EANs com faixa conferida "
          f"({sum(1 for v in web.values() if v[3])} marcados MANTER)")
    a = acoes(df, pmc, carregar_markup(), web)

    with pd.ExcelWriter(saida, engine="openpyxl") as xl:
        for nome, prio in (("1 - Faca hoje", 1), ("2 - Esta semana", 2), ("3 - Quando der", 3)):
            a[a["prioridade"].eq(prio)].to_excel(xl, sheet_name=nome, index=False)
        a.to_excel(xl, sheet_name="Tudo", index=False)
    formatar(saida)

    print(a.groupby(["prioridade", "acao"])
          .agg(itens=("ean", "count"), dinheiro=("dinheiro_em_jogo", "sum"))
          .round(0).to_string())
    print(f"\nTotal: {len(a)} acoes | Gravado: {saida}")


if __name__ == "__main__":
    main()
