from collections import defaultdict
from math import isfinite
from pathlib import Path
from statistics import median
import csv

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill


ROOT = Path.cwd()
CATALOG = ROOT / "LISTA_CLASSIFICADA_COM_PRECOS_UNITARIOS.xlsx"
REPORT = next(ROOT.glob("*com_custo_unitario.xlsx"))
CLAUDE = Path(r"C:\Users\docze\Downloads\Precificacao_Perfumaria_20260726.xlsx")
POLICY = Path(r"C:\Users\docze\.codex\skills\drogaria-precificacao\references\politica-markup-categorias.csv")
OUTPUT = ROOT / "PRECIFICACAO_COMPLETA_20260726.xlsx"
FLOOR_DIVISOR = 1 - 0.025 - 0.0598 - 0.175
GRADE = (0.49, 0.79, 0.90, 0.95, 0.99)


def ean(value):
    return "".join(char for char in str(value or "") if char.isdigit())


def currency(value):
    return None if value is None else round(value, 2)


def grid_price(target, floor):
    start, end = int(floor), int(max(target, floor)) + 2
    options = [whole + ending for whole in range(start, end + 1) for ending in GRADE if whole + ending >= floor]
    return min(options, key=lambda value: abs(value - target))


def read_policy():
    policy = {}
    with POLICY.open(encoding="utf-8-sig", newline="") as file:
        for row in csv.DictReader(file, delimiter=";"):
            try:
                policy[row["categoria_exata"]] = {
                    "role": row["papel"], "profit": float(row["lucro_liquido_alvo_pct"].replace(",", ".")) / 100,
                    "factor": float(row["fator_fisico_sobre_mediana_web"].replace(",", ".")),
                }
            except (KeyError, ValueError, AttributeError):
                continue
    # The revised taxonomy includes this missing leaf. It follows the generic equivalent
    # until a separate commercial rule is approved.
    policy["SIMILAR > ANTICONCEPCIONAL"] = {"role": "PRECO_IMAGEM", "profit": 0.10, "factor": 1.01}
    return policy


def validated_costs():
    costs = defaultdict(list)
    ws = load_workbook(REPORT, read_only=True, data_only=True).active
    for row in ws.iter_rows(min_row=6, values_only=True):
        code = ean(row[9])
        if not code or not row[10]:
            continue
        try:
            quantity, stored_cost, total = float(row[13]), float(row[25]), float(row[26])
        except (TypeError, ValueError):
            continue
        if not all(isfinite(value) and value > 0 for value in (quantity, stored_cost, total)):
            continue
        if abs(stored_cost * quantity - total) <= max(0.02, total * 0.000001):
            costs[code].append(total / quantity)
    return costs


def competitor_medians():
    medians = {}
    ws = load_workbook(CLAUDE, read_only=True, data_only=True)["Precificação Completa"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        code, value, count = ean(row[0]), row[14], row[15]
        try:
            value, count = float(value), int(count)
        except (TypeError, ValueError):
            continue
        if code and value > 0 and count > 0:
            medians[code] = (value, count)
    return medians


def main():
    policy, costs, competitors = read_policy(), validated_costs(), competitor_medians()
    catalog = load_workbook(CATALOG, read_only=True, data_only=True).active
    products = list(catalog.iter_rows(min_row=2, values_only=True))

    wb = Workbook(); full = wb.active; full.title = "Precificação Completa"
    headers = ["EAN", "Descrição", "Classificação", "Custo unitário validado", "Compras NF", "Custo mínimo NF", "Custo máximo NF", "Preço mínimo técnico", "Máximo recomendado", "Mediana concorrentes", "Nº concorrentes", "Tier", "Preço sugerido", "Status", "Justificativa"]
    full.append(headers)
    manual_rows, medicine_rows, summary = [], [], defaultdict(lambda: {"items": 0, "costed": 0, "review": 0, "cost": 0, "floor": [], "maximum": []})

    for code, description, _old_cost, classification in products:
        code = ean(code); values = costs.get(code, []); rule = policy.get(classification)
        record = summary[classification]; record["items"] += 1
        count = len(values); cost = sum(values) / count if count else None
        floor = cost / FLOOR_DIVISOR if cost else None
        ceiling = cost / (FLOOR_DIVISOR - rule["profit"]) if cost and rule else None
        market, market_count = competitors.get(code, (None, 0))
        tier = rule["role"] if rule else None
        status, suggestion, reason = "OK", None, "Preço econômico da classificação; custo validado por NF."

        if not cost:
            status, reason = "REVISAO_MANUAL_SEM_CUSTO_VALIDADO", "Sem linha de compra válida no relatório de notas."
        elif not rule:
            status, reason = "REVISAO_MANUAL_SEM_MARKUP", "Classificação sem regra financeira cadastrada."
        elif tier == "REVISAO_HUMANA":
            status, reason = "REVISAO_MANUAL", "Classificação exige conferência humana antes de definir preço."
        else:
            target = ceiling
            if market:
                physical_market = market * rule["factor"]
                if tier in {"PRECO_IMAGEM", "PADRAO"}:
                    target, reason = physical_market * 0.99, "Alvo: mediana concorrente ajustada ao canal físico, menos 1%."
                else:
                    target, reason = min(ceiling, physical_market * 1.15), "Proteção de margem, limitada quando há concorrência comparável."
            if target < floor:
                status, reason = "REVISAO_MANUAL_PISO_ACIMA_DO_MERCADO", "O mercado comparável ficou abaixo do preço mínimo técnico."
            else:
                suggestion = grid_price(target, floor)

        if classification.startswith(("ETICOS", "GENERICO", "SIMILAR")):
            if status == "OK":
                status = "REVISAO_MANUAL_TETO_CMED"
                reason = "Preço calculado sem teto CMED oficial; validar apresentação e PMC antes de aplicar."
                suggestion = None
            medicine_rows.append([code, description, classification, currency(cost), currency(floor), currency(ceiling), status])
        if status != "OK":
            manual_rows.append([code, description, classification, currency(cost), currency(floor), currency(ceiling), market, status, reason])
            record["review"] += 1
        if cost:
            record["costed"] += 1; record["cost"] += cost; record["floor"].append(floor); record["maximum"].append(ceiling)
        full.append([code, description, classification, currency(cost), count, currency(min(values)) if values else None, currency(max(values)) if values else None, currency(floor), currency(ceiling), currency(market), market_count or None, tier, currency(suggestion), status, reason])

    review = wb.create_sheet("Revisão Manual")
    review.append(["EAN", "Descrição", "Classificação", "Custo", "Preço mínimo", "Máximo recomendado", "Mediana concorrentes", "Status", "Motivo"])
    for row in manual_rows: review.append(row)
    cmed = wb.create_sheet("Medicam. sem Teto")
    cmed.append(["EAN", "Descrição", "Classificação", "Custo", "Preço mínimo", "Máximo recomendado", "Status"])
    for row in medicine_rows: cmed.append(row)

    resume = wb.create_sheet("Resumo", 0)
    resume.append(["Classificação", "Produtos", "Com custo validado", "Em revisão", "Custo médio", "Menor preço mínimo", "Maior preço mínimo", "Menor máximo recomendado", "Maior máximo recomendado"])
    for classification in sorted(summary):
        item = summary[classification]
        resume.append([classification, item["items"], item["costed"], item["review"], currency(item["cost"] / item["costed"]) if item["costed"] else None, currency(min(item["floor"])) if item["floor"] else None, currency(max(item["floor"])) if item["floor"] else None, currency(min(item["maximum"])) if item["maximum"] else None, currency(max(item["maximum"])) if item["maximum"] else None])

    notes = wb.create_sheet("Notas e Premissas")
    for line in [
        "Custos: média por EAN de linhas NF validadas por Valor Total do Item ÷ Qtde. Unitária.",
        "Preço mínimo técnico: custo ÷ 0,7402 (cartão 2,50%, Simples estimado 5,98%, despesas fixas 17,50%, lucro 0%).",
        "Máximo recomendado: alvo econômico da classificação; não é teto regulatório/CMED.",
        "Medianas concorrenciais: reaproveitadas somente por EAN idêntico do arquivo Precificacao_Perfumaria_20260726.xlsx.",
        "Medicamentos ficam em revisão até confirmação de PMC/CMED por apresentação e EAN.",
        "SIMILAR > ANTICONCEPCIONAL usa provisoriamente os mesmos parâmetros de GENERICO > ANTICONCEPCIONAL.",
    ]: notes.append([line])

    for sheet in wb.worksheets:
        sheet.freeze_panes = "A2"; sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]: cell.font = Font(bold=True, color="FFFFFF"); cell.fill = PatternFill("solid", fgColor="1F4E78")
        for column in sheet.columns:
            letter = column[0].column_letter
            sheet.column_dimensions[letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 52)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if "preço" in str(sheet.cell(1, cell.column).value).lower() or "custo" in str(sheet.cell(1, cell.column).value).lower() or "mediana" in str(sheet.cell(1, cell.column).value).lower() or "máximo" in str(sheet.cell(1, cell.column).value).lower(): cell.number_format = 'R$ #,##0.00'
        if sheet.max_column >= 1:
            for cell in sheet["A"][1:]: cell.number_format = "@"
    wb.save(OUTPUT)
    assert full.max_row - 1 == 2437
    assert all(row[3] is None or 0 < row[3] < 1000 for row in full.iter_rows(min_row=2, values_only=True))
    print(OUTPUT.resolve())
    print(f"Produtos: {full.max_row - 1}; revisão: {len(manual_rows)}; medicamentos sem teto: {len(medicine_rows)}")


if __name__ == "__main__": main()
