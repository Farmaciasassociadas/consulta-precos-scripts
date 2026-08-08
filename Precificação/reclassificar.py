#!/usr/bin/env python3
"""
Reclassify all 3,041 products in the pharmacy pricing database.
- Assign EIXO (main category) and subcategory to every product.
- Write to subcategoria_classificada table (INSERT OR REPLACE).
- Update recomendacao table for rodada 15.
- Update PRECIFICACAO_RODADA_10.xlsx with new classifications and descriptions.
"""

import sqlite3
import re
import os
from collections import Counter
import openpyxl

DB_PATH = os.path.join(os.path.dirname(__file__), 'precificador', 'precificador.db')
EXCEL_PATH = os.path.join(os.path.dirname(__file__), 'PRECIFICACAO_RODADA_10.xlsx')
LOOKUP_PATH = os.path.join(os.path.dirname(__file__), 'outputs', 'eans_pmc', 'ean_descricao_fabricante_pmc.xlsx')

FONTE = 'reclassificacao_kun_v1'

# ── Keyword definitions ──────────────────────────────────────────────

# Therapeutic subcategory keywords (for GENERICO/ETICOS/SIMILAR)
USO_CONTINUO_KW = [
    'losartan', 'losartana', 'enalapril', 'sinvastatina', 'metformina',
    'levotiroxina', 'omeprazol', 'fluoxetina', 'sertralina',
    'hidroclorotiazida', 'anlodipino', 'atenolol', 'captopril',
    'glibenclamida', 'gliclazida', 'insulina', 'citalopram',
    'amitriptilina', 'nortriptilina', 'carbamazepina', 'fenitoina',
    'fenobarbital', 'acido valproico', 'valproato', 'lamotrigina',
    'topiramato', 'pregabalina', 'gabapentina', 'duloxetina',
    'venlafaxina', 'paroxetina', 'escitalopram', 'bupropiona',
    'risperidona', 'olanzapina', 'quetiapina', 'aripiprazol',
    'haloperidol', 'clorpromazina', 'levodopa', 'pramipexol',
    'bromocriptina', 'biperideno', 'donepezila', 'galantamina',
    'rivastigmina', 'memantina', 'alopurinol', 'colchicina',
    'propranolol', 'carvedilol', 'metoprolol', 'bisoprolol',
    'espironolactona', 'furosemida', 'hidralazina', 'isossorbida',
    'digoxina', 'amiodarona', 'varfarina', 'warfarin', 'acenocumarol',
    'rivaroxabana', 'apixabana', 'dabigatrana', 'aas', 'acido acetilsalicilico',
    'clopidogrel', 'ticagrelor', 'rosuvastatina', 'atorvastatina',
    'bezafibrato', 'ciprofibrato', 'fenofibrato', 'ezetimiba',
    'pantoprazol', 'esomeprazol', 'lansoprazol', 'ranitidina',
    'budesonida', 'formoterol', 'salmeterol', 'fluticasona',
    'brometo de tiotropio', 'tiotropio', 'ipratropio',
    'montelucaste', 'levocetirizina', 'desloratadina',
    'prednisona', 'prednisolona', 'dexametasona',
    'metotrexato', 'azatioprina', 'ciclosporina', 'tacrolimo',
    'micofenolato', 'sulfassalazina', 'hidroxicloroquina',
    'sildenafil', 'tadalafil', 'vardenafil',
    'finasterida', 'dutasterida', 'tansulosina', 'doxazosina',
    'alendronato', 'risedronato', 'calcio', 'vitamina d',
    'colecalciferol', 'carbonato de calcio', 'citrato de calcio',
    'repositor', 'hormonal', 'estradiol', 'estriol', 'progesterona',
    'tireoide', 'propiltiouracila', 'metimazol',
    'mesalazina', 'mesalamina', 'infliximabe', 'adalimumabe',
]

ANTIMICROBIANO_KW = [
    'amoxicilina', 'azitromicina', 'ciprofloxacino', 'cefalexina',
    'ceftriaxona', 'cefuroxima', 'cefaclor', 'cefadroxil',
    'claritromicina', 'eritromicina', 'tetraciclina', 'doxiciclina',
    'minociclina', 'sulfametoxazol', 'trimetoprima', 'bactrim',
    'metronidazol', 'secnidazol', 'tinidazol', 'nitrofurantoina',
    'norfloxacino', 'levofloxacino', 'moxifloxacino', 'ofloxacino',
    'gentamicina', 'neomicina', 'amicacina', 'clindamicina',
    'lincomicina', 'penicilina', 'benzilpenicilina', 'benzetacil',
    'ampicilina', 'amoxilina', 'fluconazol', 'itraconazol',
    'cetoconazol', 'miconazol', 'nistatina', 'terbinafina',
    'aciclovir', 'valaciclovir', 'oseltamivir', 'ivermectina',
    'albendazol', 'mebendazol', 'praziquantel', 'cloroquina',
    'primaquina', 'artesunato', 'artemeter', 'lumefantrina',
    'rifampicina', 'isoniazida', 'pirazinamida', 'etambutol',
    'dapsona', 'anfotericina', 'vancomicina', 'teicoplanina',
    'meropenem', 'imipenem', 'ertapenem', 'piperacilina',
    'tazobactam', 'sulbactam', 'acido clavulanico', 'clavulanato',
    'fosfomicina', 'nitazoxanida', 'tinidazol',
]

OTC_MIP_KW = [
    'paracetamol', 'dipirona', 'ibuprofeno', 'loratadina',
    'cetirizina', 'fexofenadina', 'desloratadina',
    'acido acetilsalicilico', 'aas', 'aspirina',
    'naproxeno', 'cetoprofeno', 'diclofenaco', 'nimesulida',
    'buscopan', 'escopolamina', 'butilbrometo',
    'simeticona', 'dimeticona', 'luftal',
    'antiacido', 'hidroxido de aluminio', 'hidroxido de magnesio',
    'omeprazol', 'pantoprazol',
    'ranitidina', 'cimetidina',
    'vitamina c', 'vit c', 'acido ascorbico',
    'complexo b', 'vitamina b',
    'multivitaminico', 'polivitaminico',
    'sais minerais', 'soro', 'reidratante', 'eletrolito',
    'colageno', 'colageno hidrolisado',
    'glicosamina', 'condroitina',
    'propolis', 'equinacea', 'melatonina',
    'laxante', 'bisacodil', 'lactulose', 'lactulona',
    'gripe', 'antigripal', 'resfriado',
    'descongestionante', 'nasal',
    'pastilha', 'garganta', 'tosse', 'expectorante',
    'pomada', 'gel', 'antiinflamatorio', 'antiflamatorio',
    'agua oxigenada', 'alcool', 'algodao', 'esparadrapo',
    'termometro', 'seringa', 'agulha', 'lanceta',
    'curativo', 'band aid', 'band-aid', 'atadura', 'gaze',
    'alcool gel', 'alcool em gel',
]

CONTROLADO_KW = [
    'clonazepam', 'alprazolam', 'codeina', 'codein', 'tramadol',
    'diazepam', 'lorazepam', 'bromazepam', 'midazolam',
    'flunitrazepam', 'nitrazepam', 'clobazam', 'cloxazolam',
    'zolpidem', 'zopiclona', 'eszopiclona',
    'metilfenidato', 'ritalina', 'concerta', 'venvanse', 'lisdexanfetamina',
    'morfina', 'oxicodona', 'oxicodon', 'metadona', 'fentanil', 'fentanila',
    'buprenorfina', 'naloxona', 'naltrexona',
    'fenobarbital', 'primidona', 'etossuximida',
    'clordiazepoxido', 'meprobamato',
    'anfepramona', 'femproporex', 'mazindol', 'sibutramina',
    'testosterona', 'undecilato', 'cipionato',
    'hemitartarato', 'zolpidem',
    'pregabalina', 'gabapentina',
    'ciclobenzaprina', 'carisoprodol',
    'ropinirol', 'selegilina', 'entacapona',
    'clozapina', 'paliperidona', 'ziprasidona',
    'metilfenidato', 'dexanfetamina',
    'petidina', 'meperidina',
]

ANTICONCEPCIONAL_KW = [
    'levonorgestrel', 'etinilestradiol', 'desogestrel', 'gestodeno',
    'noretisterona', 'noretindrona', 'acetato de medroxiprogesterona',
    'medroxiprogesterona', 'drospirenona', 'ciproterona',
    'dienogest', 'norgestimato', 'elcometrina', 'ulipristal',
    'anticoncepcional', 'contraceptivo', 'pílula', 'pilula',
    'dIU', 'diu ', ' diu', 'implante', 'implanon',
    'adstringente', 'espermicida',
    'nuvaring', 'anel vaginal',
    'mesigyna', 'ciclo 21', 'microvlar', 'tamisa',
    'diane 35', 'selene', 'yasmin', 'yaz', 'cerazette',
    'injecao', 'injetavel', 'trimestral', 'mensal',
]

# Perfumaria subcategories
PERFUMARIA_KW = {
    'TINTURAS': [
        'tintura', 'coloracao', 'coloração', 'tinta', 'water color',
        'agua oxig', 'água oxig', 'oxidante', 'descolorante',
        'po descolorante', 'pó descolorante', 'luzes', 'reflexo',
        'tonalizante', 'revelador', 'creme revelador',
    ],
    'FRALDAS': [
        'fralda', 'fraldinha', 'cueiro', 'calça fralda',
        'lenço umedecido', 'lenco umedecido', 'hastes flexiveis',
        'pomada para assadura', 'pomada assadura', 'hipoglos', 'hipoglós',
    ],
    'HIGIENE BUCAL': [
        'creme dental', 'pasta dental', 'pasta de dente', 'dentifricio',
        'escova dental', 'escova de dente', 'escova dente',
        'fio dental', 'fita dental', 'palito dental',
        'enxaguatorio', 'enxaguante', 'antisseptico bucal',
        'limpador de protese', 'protese dentaria',
        'fixador de protese', 'fixodent', 'corega',
        'cera orto', 'orto', 'gum ', 'gum-',
    ],
    'MAQUIAGEM': [
        'maquiagem', 'maquilagem', 'base ', 'corretivo',
        'rimel', 'rímel', 'mascara cilios', 'máscara cílios',
        'lapis olho', 'lápis olho', 'delineador', 'sombra ',
        'blush', 'po compacto', 'pó compacto', 'po facial',
        'batom', 'gloss', 'lapis boca', 'lápis boca',
        'primer', 'iluminador', 'contorno', 'paleta',
        'cilio postico', 'cílio postiço', 'the creme',
        'esmalte', 'acetona', 'removedor esmalte', 'lixa unha',
        'base unha', 'extra brilho', 'oleo secante',
        'pinça', 'pinça sobrancelha', 'curvex',
    ],
    'LINHA INFANTIL': [
        'infantil', 'bebe', 'bebê', 'baby', 'kids', 'kid ',
        'crianca', 'criança', 'nursery', 'nenem', 'neném',
        'mamadeira', 'chupeta', 'mordedor',
        'shampoo infantil', 'shampoo bebe', 'shampoo bebê',
        'condicionador infantil', 'sabonete infantil', 'sabonete bebe',
        'colonia infantil', 'colônia infantil',
    ],
    'DERMOCOSMETICOS': [
        'dermocosmetico', 'dermocosmético',
        'acido hialuronico', 'ácido hialurônico', 'hialuronico',
        'retinol', 'niacinamida', 'vitamina c', 'vit c',
        'serum', 'sérum', 'anti-idade', 'antiidade', 'antienvelhecimento',
        'anti rugas', 'antirrugas', 'antissinais', 'anti sinais',
        'clareador', 'clareamento', 'uniformizador', 'renovador',
        'peeling', 'esfoliante', 'esfoliacao',
        'protecao solar', 'proteção solar', 'protetor solar', 'prot solar',
        'fotoprotetor', 'bloqueador solar', 'fps ',
        'hidratante facial', 'hidratante anti',
        'gel de limpeza', 'gel limp', 'gel limpante',
        'agua micelar', 'água micelar', 'micelar',
        'tonico facial', 'tônico facial',
        'loção hidratante', 'loção corporal',
        'emulsao', 'emulsão', 'creme hidratante',
        'principia', 'eucerin', 'la roche', 'vichy', 'avene', 'avène',
        'bioderma', 'nivea', 'neutrogena', 'cerave', 'cetaphil',
        'creme para maos', 'creme para mãos', 'creme pés',
    ],
    'LINHA CAPILAR': [
        'shampoo', 'condicionador', 'mascara capilar', 'máscara capilar',
        'creme de pentear', 'leave-in', 'leave in', 'finalizador',
        'oleo capilar', 'óleo capilar', 'reparador pontas',
        'reconstrutor', 'queratina', 'cauterizacao',
        'alisante', 'relaxante', 'progressiva',
        'gel capilar', 'gel modelador', 'fixador cabelo',
        'spray capilar', 'laquê', 'laque',
        'coloracao', 'coloração',
    ],
    'BRONZ & HIDRATANTES': [
        'bronzeador', 'bronz', 'protetor solar', 'bloqueador solar',
        'hidratante', 'loção', 'loção corporal', 'loção hidratante',
        'creme corporal', 'creme para corpo', 'creme p corpo',
        'oleo corporal', 'óleo corporal', 'óleo de amendoas',
        'manteiga corporal', 'manteiga de karite',
        'pos sol', 'pós sol', 'after sun', 'calmante',
        'repelente', 'inseticida', 'off ',
        'paixao oleo', 'paixão óleo', 'paixao loção', 'paixao hidr',
    ],
    'COLONIAS & DESODORANTES': [
        'desodorante', 'desodorante', 'antitranspirante', 'antiperspirante',
        'rexona', 'dove des', 'nivea des', 'rexona clinical',
        'colonia', 'colônia', 'perfume', 'eau de', 'deo parfum',
        'sabonete', 'sabonete liquido', 'sabonete em barra',
        'sabonete intimo', 'sabonete íntimo',
    ],
    'HIGIENE PESSOAL': [
        'nivea', 'dove', 'johnson', 'natura', 'phebo', 'granado',
        'absorvente', 'protetor diario', 'protetor diário',
        'papel higienico', 'papel higiênico',
        'algodao', 'algodão', 'cotonete', 'hastes flexiveis',
        'lamina', 'lâmina', 'aparelho barbear', 'prestobarba',
        'creme barbear', 'espuma barbear', 'gel barbear', 'pos barba', 'pós barba',
        'sabonete', 'saboneteira', 'sabonete liquido', 'sabonete em barra',
        'desodorante', 'antitranspirante',
        'talco', 'talquinho',
        'escova cabelo', 'pente', 'escova modelador',
        'depilacao', 'depilação', 'cera depilatoria',
    ],
}

# Varejo subcategories
VAREJO_KW = {
    'BEBIDAS E BOMBONIERE': [
        'bebida', 'refrigerante', 'coca cola', 'coca-cola', 'pepsi',
        'suco', 'cha', 'chá', 'energetico', 'energético', 'monster',
        'agua mineral', 'água mineral', 'agua de coco', 'água de coco',
        'isotônico', 'isotónico', 'gatorade',
        'bombom', 'chocolate', 'barra cereal', 'biscoito',
        'doce', 'balinha', 'balas', 'caramelo', 'guloseima',
        'mentos', 'trident', 'halls', 'tictac',
        'garoto', 'lacta', 'nestle', 'nestlé', 'sonho de valsa',
        'snack', 'salgadinho', 'amendoim', 'castanha',
    ],
    'SUPLEMENTOS': [
        'suplemento', 'whey', 'proteina', 'proteína',
        'bcaa', 'creatina', 'glutamina', 'arginina', 'beta alanina',
        'pre treino', 'pré treino', 'pre-treino', 'termogenico',
        'hipercalorico', 'hipercalórico', 'maltodextrina',
        'albumina', 'caseina', 'isolada', 'hidrolisada',
        'santo habito', 'revigore',
        'colageno', 'colágeno', 'omega 3', 'ômega 3',
        'coenzima', 'coenzima q10', 'q10',
        'magnesio', 'magnésio', 'zinco', 'selenio', 'selênio',
        'cromo', 'ferro', 'potassio', 'potássio',
        'vitamina', 'vit ', 'vit.',
        'aminoacido', 'aminoácido',
        'barra proteina', 'barra de proteina',
        'energy 3d', 'energy 5g',
        'memoria', 'memória', 'libid', 'libido',
        'triacalm', 'metaslim', 'immunity', 'inmunity',
        'digest', 'condro', 'cranberry', 'cafeina',
        'acido hialuronico', 'colageno',
        'az homem', 'az mulher', 'az 34', '50+',
        'beauty', 'feno grego', 'femme',
        'curcuma', 'cúrcuma', 'spirulina', 'chlorella',
    ],
    'FITOTERAPICOS': [
        'fitoterapico', 'fitoterápico', 'herbal', 'natural',
        'extrato seco', 'extrato fluido', 'extrato glicolico',
        'tintura', 'alcoolatura',
        'arnica', 'calendula', 'camomila', 'erva doce',
        'erva cidreira', 'boldo', 'carqueja', 'alcachofra',
        'guaco', 'guaraná', 'guarana', 'gengibre',
        'ginkgo', 'ginseng', 'valeriana', 'maracuja', 'maracujá',
        'passiflora', 'melissa', 'hipérico', 'hiperico',
        'castanha da india', 'centella', 'confrei',
        'alcachofra', 'alcaxofra', 'sene', 'cáscara sagrada',
        'espinheira santa', 'unha de gato', 'garra do diabo',
        'saw palmetto', 'pygeum', 'catuaba', 'muira puama',
        'maca peruana', 'tribulus', 'long jack',
    ],
    'LEITES NUTRICAO': [
        'nutricao', 'nutrição', 'nutren', 'ensure', 'forti',
        'nutridrink', 'nutrison', 'diet', 'dieta', 'enteral',
        'suplemento nutricional', 'alimentacao', 'alimentação',
        'proteina', 'hipercalorico', 'hiperproteico',
        'fiber', 'fibra', 'glicerna', 'glucerna',
        'pediasure', 'nutren', 'nutri', 'milkshake',
        'diamax', 'isosource', 'novasource',
    ],
    'LEITES': [
        'leite em po', 'leite em pó', 'leite po', 'leite pó',
        'leite infantil', 'leite ninho', 'leite nestle',
        'formula infantil', 'fórmula infantil',
        'nan ', 'nestogeno', 'aptamil', 'enfamil',
        'milupa', 'nestle', 'ninho',
        'leite integral', 'leite desnatado', 'leite semidesnatado',
        'leite zero', 'leite sem lactose',
        'composto lacteo', 'composto lácteo',
        'farinha lactea', 'farinha láctea', 'mucilon',
    ],
    'ORTOPEDICOS': [
        'ortopedico', 'ortopédico', 'ortopedia',
        'cinta', 'joelheira', 'tornozeleira', 'munhequeira',
        'cotoveleira', 'imobilizador', 'tala',
        'meia elastica', 'meia elástica', 'meia compressao',
        'bolsa agua quente', 'bolsa termica', 'bolsa térmica',
        'compressa', 'gel quente', 'gel frio',
        'almofada', 'travesseiro ortopedico', 'colchao',
        'palmilha', 'calcanheira', 'protetor calcanhar',
        'tornozeleira', 'faixa elastica', 'faixa elástica',
        'tipoia', 'tipóia', 'muleta', 'bengala',
        'bolsa agua', 'bioclima', 'bioland',
    ],
    'CONVENIENCIAS': [
        'pilha', 'bateria', 'carregador', 'cabo',
        'caneta', 'lapis', 'lápis', 'borracha', 'caderno',
        'tesoura', 'fita adesiva', 'durex', 'cola',
        'isqueiro', 'fosforo', 'fósforo',
        'chaveiro', 'guarda chuva', 'capa chuva',
        'oculos', 'óculos', 'leitura',
        'preservativo', 'camisinha', 'lubrificante',
        'gel lubrificante', 'oleo massage', 'óleo massage',
    ],
    'ACESSORIOS': [
        'acessorio', 'acessório',
        'copo', 'garrafa', 'squeeze', 'mamadeira',
        'organizador', 'porta', 'estojo', 'necessaire',
        'espelho', 'pente', 'escova modelador',
        'bolsa termica', 'sacola termica',
        'nebulizador', 'inalador', 'inalador',
        'termometro', 'medidor pressao', 'medidor pressão',
        'oximetro', 'oxímetro',
        'tira', 'tiras', 'lanceta', 'lancetas',
        'seringa', 'agulha', 'coletor',
        'inalador', 'nebulizador', 'aerocamara',
        'frasco', 'conta gotas', 'dosador',
    ],
    'VITAMINAS': [
        'vitamina', 'polivitaminico', 'multivitaminico',
        'complexo vitaminico', 'suplemento vitaminico',
        'vit a', 'vit b', 'vit c', 'vit d', 'vit e', 'vit k',
        'acido folico', 'ácido fólico', 'biotina',
        'riboflavina', 'tiamina', 'piridoxina', 'cianocobalamina',
        'retinol', 'tocoferol', 'filoquinona',
        'revit', 'centrum', 'gerovital', 'vitergan', 'viterra',
        'calcio', 'cálcio', 'ferro', 'zinco',
        'suplemento mineral',
    ],
}

# Brand-specific overrides
BRAND_EIXO = {
    'santo habito': 'VAREJO',
    'revigore': 'VAREJO',
    'principia': 'PERFUMARIA',
    'eucerin': 'PERFUMARIA',
    'rexona': 'PERFUMARIA',
    'nivea': 'PERFUMARIA',
    'dove': 'PERFUMARIA',
}

BRAND_SUBCAT = {
    'santo habito': 'SUPLEMENTOS',
    'revigore': 'SUPLEMENTOS',
    'principia': 'DERMOCOSMETICOS',
    'eucerin': 'DERMOCOSMETICOS',
    'rexona': 'COLONIAS & DESODORANTES',
    'nivea': 'HIGIENE PESSOAL',
    'dove': 'HIGIENE PESSOAL',
}

# ── Helper functions ──────────────────────────────────────────────────

def normalize(text):
    """Normalize text for keyword matching."""
    if text is None:
        return ''
    return text.lower().strip()


def matches_any(text, keywords):
    """Check if any keyword appears in normalized text."""
    t = normalize(text)
    for kw in keywords:
        if normalize(kw) in t:
            return True
    return False


def get_brand(descricao):
    """Extract brand name from description."""
    d = normalize(descricao)
    for brand in BRAND_EIXO:
        if brand in d:
            return brand
    return None


def classify_eixo(descricao, grupo_pai_nf, segmento):
    """Determine EIXO (main category)."""
    gp = normalize(grupo_pai_nf)
    seg = normalize(segmento)
    desc = normalize(descricao)

    # Brand override
    brand = get_brand(descricao)
    if brand and brand in BRAND_EIXO:
        return BRAND_EIXO[brand]

    # Brick segment
    if seg == 'gen':
        return 'GENERICO'
    if seg == 'rx':
        return 'ETICOS'
    if seg == 'sim':
        return 'SIMILAR'

    # NF group
    if 'genérico' in gp or 'generico' in gp:
        return 'GENERICO'
    if 'referência' in gp or 'referencia' in gp:
        return 'ETICOS'
    if 'similar' in gp:
        return 'SIMILAR'
    if 'perfumaria' in gp:
        return 'PERFUMARIA'

    # LIBERADO or no data -> description analysis
    # Check for perfume/cosmetic keywords
    perfume_kw = ['perfumaria', 'sabonete', 'shampoo', 'condicionador', 'desodorante',
                  'maquiagem', 'tintura', 'creme dental', 'fralda', 'protetor solar',
                  'hidratante', 'colonia', 'bronzeador', 'esmalte', 'batom',
                  'absorvente', 'cotonete', 'lamina barbear', 'prestobarba',
                  'desodorante', 'antitranspirante', 'perfume']
    if matches_any(desc, perfume_kw):
        return 'PERFUMARIA'

    # Check for varejo keywords
    varejo_kw = ['suplemento', 'whey', 'proteina', 'creatina', 'bcaa',
                 'vitamina', 'polivitaminico', 'bebida', 'refrigerante',
                 'energetico', 'bombom', 'chocolate', 'leite em pó', 'leite po',
                 'fralda', 'ortopedico', 'cinta', 'joelheira',
                 'pilha', 'bateria', 'caneta', 'preservativo',
                 'barra cereal', 'biscoito', 'suco', 'cha ',
                 'acessorio', 'copo', 'garrafa', 'squeeze',
                 'oleo corporal', 'loção', 'creme corporal',
                 'agua oxig', 'pedra hume',
                 'santo habito', 'revigore']
    if matches_any(desc, varejo_kw):
        return 'VAREJO'

    # Check for generic/eticos/similar keywords in description
    drug_kw = ['mg ', 'mg/', 'mcg', 'mcg/', '% ', 'com', 'cap', 'cps', 'comp',
               'sol inj', 'sol oral', 'susp', 'xpe', 'xarope', 'gotas',
               'pomada', 'creme', 'gel ', 'comprimido', 'capsula', 'cápsula',
               'injetavel', 'injetável', 'ampola', 'frasco ampola']
    if matches_any(desc, drug_kw):
        # Default to ETICOS for unrecognized drugs
        return 'ETICOS'

    # Default fallback
    return 'VAREJO'


def classify_subcategoria(descricao, grupo_filho_nf, eixo):
    """Determine subcategory within EIXO."""
    gf = normalize(grupo_filho_nf)
    desc = normalize(descricao)

    # Brand subcategory override
    brand = get_brand(descricao)
    if brand and brand in BRAND_SUBCAT:
        sub = BRAND_SUBCAT[brand]
        if eixo == 'VAREJO' and sub == 'SUPLEMENTOS':
            return 'VAREJO > SUPLEMENTOS'
        if eixo == 'PERFUMARIA' and sub in ['DERMOCOSMETICOS', 'COLONIAS & DESODORANTES', 'HIGIENE PESSOAL']:
            return f'PERFUMARIA > {sub}'

    if eixo in ('GENERICO', 'ETICOS', 'SIMILAR'):
        return classify_therapeutic(desc, gf, eixo)
    elif eixo == 'PERFUMARIA':
        return classify_perfumaria(desc, gf)
    elif eixo == 'VAREJO':
        return classify_varejo(desc, gf)
    else:
        return eixo  # fallback


def classify_therapeutic(desc, grupo_filho_nf, eixo):
    """Classify therapeutic subcategory (GENERICO/ETICOS/SIMILAR)."""
    gf = normalize(grupo_filho_nf)

    # Use grupo_filho_nf hints first
    gf_hints = {
        'genérico controlado': 'CONTROLADO',
        'generico controlado': 'CONTROLADO',
        'referencia controlado': 'CONTROLADO',
        'generico antimicrobiano': 'ANTIMICROBIANO',
        'referencia antimicrobiano': 'ANTIMICROBIANO',
        'genérico anticoncepcional': 'ANTICONCEPCIONAL',
        'generico anticoncepcional': 'ANTICONCEPCIONAL',
        'referencia anticoncepcional': 'ANTICONCEPCIONAL',
        'similar anticoncepcional': 'ANTICONCEPCIONAL',
        'generico fcia popular': 'USO CONTINUO',
        'referencia fcia popular': 'USO CONTINUO',
        'similar fcia popular': 'USO CONTINUO',
    }
    for hint_key, sub in gf_hints.items():
        if hint_key in gf:
            # Adjust subcategory prefix for SIMILAR
            if eixo == 'SIMILAR':
                if sub == 'O.T.C/MIP':
                    return 'SIMILAR > O.T.C/MIP-SIMILAR'
                if sub == 'RX':
                    return 'SIMILAR > RX-SIMILAR'
            return f'{eixo} > {sub}'

    # Check description keywords
    # Order matters: check specific before general
    if matches_any(desc, ANTICONCEPCIONAL_KW):
        return f'{eixo} > ANTICONCEPCIONAL'
    if matches_any(desc, CONTROLADO_KW):
        return f'{eixo} > CONTROLADO'
    if matches_any(desc, ANTIMICROBIANO_KW):
        return f'{eixo} > ANTIMICROBIANO'
    if matches_any(desc, USO_CONTINUO_KW):
        return f'{eixo} > USO CONTINUO'
    if matches_any(desc, OTC_MIP_KW):
        if eixo == 'SIMILAR':
            return 'SIMILAR > O.T.C/MIP-SIMILAR'
        return f'{eixo} > O.T.C/MIP'

    # Default: RX (prescription)
    if eixo == 'SIMILAR':
        return 'SIMILAR > RX-SIMILAR'
    return f'{eixo} > RX'


def classify_perfumaria(desc, grupo_filho_nf):
    """Classify PERFUMARIA subcategory."""
    gf = normalize(grupo_filho_nf)

    # Use grupo_filho_nf
    if 'fraldas' in gf:
        return 'PERFUMARIA > FRALDAS'

    # Check in priority order
    priority_order = [
        'FRALDAS', 'HIGIENE BUCAL', 'MAQUIAGEM', 'LINHA INFANTIL',
        'TINTURAS', 'DERMOCOSMETICOS', 'LINHA CAPILAR',
        'BRONZ & HIDRATANTES', 'COLONIAS & DESODORANTES', 'HIGIENE PESSOAL',
    ]

    for sub in priority_order:
        if matches_any(desc, PERFUMARIA_KW.get(sub, [])):
            return f'PERFUMARIA > {sub}'

    return 'PERFUMARIA'


def classify_varejo(desc, grupo_filho_nf):
    """Classify VAREJO subcategory."""
    gf = normalize(grupo_filho_nf)

    if 'nutrição' in gf or 'nutricao' in gf or 'leites' in gf:
        return 'VAREJO > LEITES NUTRICAO'

    priority_order = [
        'BEBIDAS E BOMBONIERE', 'SUPLEMENTOS', 'FITOTERAPICOS',
        'LEITES NUTRICAO', 'LEITES', 'ORTOPEDICOS',
        'CONVENIENCIAS', 'ACESSORIOS', 'VITAMINAS',
    ]

    for sub in priority_order:
        if matches_any(desc, VAREJO_KW.get(sub, [])):
            return f'VAREJO > {sub}'

    return 'VAREJO'


def sanitize_ean_for_excel(ean):
    """Convert EAN to match Excel format: zero-pad to 13 chars."""
    return str(ean).zfill(13)


def main():
    print("=" * 70)
    print("RECLASSIFICAÇÃO DE PRODUTOS - Kun v1")
    print("=" * 70)

    # ── 1. Connect to DB ──────────────────────────────────────────────
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    # ── 2. Load all products ──────────────────────────────────────────
    cur.execute('SELECT ean, descricao, grupo_pai_nf, grupo_filho_nf FROM produto')
    produtos = cur.fetchall()
    print(f"\nTotal products: {len(produtos)}")

    # Load preco_brick segment mapping
    cur.execute('SELECT ean, segmento FROM preco_brick')
    brick_map = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Brick mappings: {len(brick_map)}")

    # Load existing subcategoria_classificada (to preserve)
    cur.execute("SELECT ean, classificacao_exata, fonte FROM subcategoria_classificada WHERE fonte != ?", (FONTE,))
    existing = {row[0]: row[1] for row in cur.fetchall()}
    print(f"Existing classifications (non-Kun): {len(existing)}")

    # ── 3. Classify all products ──────────────────────────────────────
    results = {}  # ean -> classificacao_exata
    stats = Counter()
    eixo_stats = Counter()

    for ean, desc, gp, gf in produtos:
        # Preserve existing manual classifications
        if ean in existing:
            results[ean] = existing[ean]
            stats['preserved_existing'] += 1
            continue

        segmento = brick_map.get(ean, '')

        eixo = classify_eixo(desc, gp, segmento)
        eixo_stats[eixo] += 1

        classificacao = classify_subcategoria(desc, gf, eixo)
        results[ean] = classificacao
        stats['newly_classified'] += 1

    print(f"\nPreserved existing: {stats['preserved_existing']}")
    print(f"Newly classified: {stats['newly_classified']}")

    print(f"\nEIXO distribution:")
    for eixo, count in eixo_stats.most_common():
        print(f"  {eixo}: {count}")

    # Subcategory distribution (new classifications)
    print(f"\nSubcategory distribution (new classifications):")
    subcat_new = Counter()
    for ean, (desc, gp, gf) in [(p[0], (p[1], p[2], p[3])) for p in produtos]:
        if ean not in existing:
            subcat_new[results[ean]] += 1
    for cat, count in subcat_new.most_common():
        print(f"  {cat}: {count}")

    # ── 4. Write to subcategoria_classificada ─────────────────────────
    write_count = 0
    for ean, cat in results.items():
        if ean in existing:
            continue  # don't overwrite existing manual ones
        cur.execute(
            'INSERT OR REPLACE INTO subcategoria_classificada (ean, classificacao_exata, fonte) VALUES (?, ?, ?)',
            (ean, cat, FONTE)
        )
        write_count += 1

    conn.commit()
    print(f"\n✓ Written {write_count} classifications to subcategoria_classificada")

    # ── 5. Update recomendacao table for rodada 15 ────────────────────
    cur.execute("SELECT ean FROM recomendacao WHERE rodada_id = 15")
    rec_eans = {row[0] for row in cur.fetchall()}

    updated_rec = 0
    for ean in rec_eans:
        ean_13 = sanitize_ean_for_excel(ean)
        # Try exact match first, then zfilled
        cat = None
        if ean in results:
            cat = results[ean]
        elif ean_13 in results:
            cat = results[ean_13]

        if cat:
            cur.execute(
                'UPDATE recomendacao SET categoria_provisoria = ? WHERE rodada_id = 15 AND ean = ?',
                (cat, ean)
            )
            updated_rec += 1

    conn.commit()
    print(f"✓ Updated {updated_rec} entries in recomendacao (rodada 15)")

    # Verify
    cur.execute("SELECT COUNT(*) FROM recomendacao WHERE rodada_id=15 AND (categoria_provisoria IS NULL OR categoria_provisoria='')")
    still_empty = cur.fetchone()[0]
    print(f"  (Still empty after update: {still_empty})")

    conn.close()

    # ── 6. Load lookup descriptions ───────────────────────────────────
    print("\nLoading description lookup from Excel...")
    wb_lookup = openpyxl.load_workbook(LOOKUP_PATH, data_only=True)
    ws_lookup = wb_lookup['Produtos']

    desc_lookup = {}
    for r in range(2, ws_lookup.max_row + 1):
        ean_raw = ws_lookup.cell(row=r, column=1).value
        desc_raw = ws_lookup.cell(row=r, column=2).value
        if ean_raw is not None and desc_raw is not None:
            ean_str = str(ean_raw).strip()
            desc_str = str(desc_raw).strip()
            # Store both raw and zfilled
            desc_lookup[ean_str] = desc_str
            desc_lookup[ean_str.zfill(13)] = desc_str

    print(f"Lookup entries: {len(set(desc_lookup.values()))}")

    # ── 7. Update Excel file ──────────────────────────────────────────
    print("\nUpdating PRECIFICACAO_RODADA_10.xlsx...")
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb['Precificação Completa']

    excel_updated_cat = 0
    excel_updated_desc = 0
    excel_eans_total = 0

    for r in range(2, ws.max_row + 1):
        excel_eans_total += 1
        ean_cell = ws.cell(row=r, column=1)
        ean_excel = str(ean_cell.value).strip() if ean_cell.value is not None else ''

        # Try to match EAN to our results
        # Excel EANs are zfilled to 13; DB eans are raw
        cat = None
        if ean_excel in results:
            cat = results[ean_excel]
        else:
            # Try stripping leading zeros
            ean_stripped = ean_excel.lstrip('0') or '0'
            if ean_stripped in results:
                cat = results[ean_stripped]

        if cat:
            ws.cell(row=r, column=3).value = cat
            excel_updated_cat += 1

        # Fill missing descriptions
        existing_desc = ws.cell(row=r, column=2).value
        if existing_desc is None or str(existing_desc).strip() == '' or str(existing_desc).strip().endswith('*'):
            # Try lookup
            desc_new = None
            if ean_excel in desc_lookup:
                desc_new = desc_lookup[ean_excel]
            else:
                ean_stripped = ean_excel.lstrip('0') or '0'
                if ean_stripped in desc_lookup:
                    desc_new = desc_lookup[ean_stripped]

            if desc_new:
                ws.cell(row=r, column=2).value = desc_new
                excel_updated_desc += 1

    print(f"  Excel rows: {excel_eans_total}")
    print(f"  Categories updated: {excel_updated_cat}")
    print(f"  Descriptions filled: {excel_updated_desc}")

    # ── 8. Add summary sheet ──────────────────────────────────────────
    if 'Resumo Reclassificação' in wb.sheetnames:
        del wb['Resumo Reclassificação']

    ws_summary = wb.create_sheet('Resumo Reclassificação')

    ws_summary.cell(row=1, column=1).value = 'Resumo da Reclassificação - Kun v1'
    ws_summary.cell(row=1, column=1).font = openpyxl.styles.Font(bold=True, size=14)

    ws_summary.cell(row=3, column=1).value = 'Total de produtos:'
    ws_summary.cell(row=3, column=2).value = len(produtos)

    ws_summary.cell(row=4, column=1).value = 'Classificações preservadas (existentes):'
    ws_summary.cell(row=4, column=2).value = stats['preserved_existing']

    ws_summary.cell(row=5, column=1).value = 'Novas classificações:'
    ws_summary.cell(row=5, column=2).value = stats['newly_classified']

    ws_summary.cell(row=7, column=1).value = 'Distribuição por EIXO (novos):'
    ws_summary.cell(row=7, column=1).font = openpyxl.styles.Font(bold=True)
    row = 8
    for eixo, count in eixo_stats.most_common():
        ws_summary.cell(row=row, column=1).value = eixo
        ws_summary.cell(row=row, column=2).value = count
        row += 1

    row += 1
    ws_summary.cell(row=row, column=1).value = 'Distribuição por subcategoria (novos):'
    ws_summary.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
    row += 1
    for cat, count in subcat_new.most_common():
        ws_summary.cell(row=row, column=1).value = cat
        ws_summary.cell(row=row, column=2).value = count
        row += 1

    row += 1
    ws_summary.cell(row=row, column=1).value = 'Atualizações:'
    ws_summary.cell(row=row, column=1).font = openpyxl.styles.Font(bold=True)
    row += 1
    ws_summary.cell(row=row, column=1).value = 'Recomendação (rodada 15):'
    ws_summary.cell(row=row, column=2).value = updated_rec
    row += 1
    ws_summary.cell(row=row, column=1).value = 'Excel categorias atualizadas:'
    ws_summary.cell(row=row, column=2).value = excel_updated_cat
    row += 1
    ws_summary.cell(row=row, column=1).value = 'Excel descrições preenchidas:'
    ws_summary.cell(row=row, column=2).value = excel_updated_desc

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 45
    ws_summary.column_dimensions['B'].width = 20

    wb.save(EXCEL_PATH)
    print(f"\n✓ Excel saved with summary sheet 'Resumo Reclassificação'")

    # ── 9. Export summary text file ──────────────────────────────────
    summary_path = os.path.join(os.path.dirname(__file__), 'reclassificacao_summary.txt')
    with open(summary_path, 'w', encoding='utf-8') as f:
        f.write("=" * 70 + "\n")
        f.write("RECLASSIFICAÇÃO DE PRODUTOS - Kun v1\n")
        f.write("=" * 70 + "\n\n")
        f.write(f"Fonte: {FONTE}\n")
        f.write(f"Total de produtos: {len(produtos)}\n")
        f.write(f"Classificações preservadas (existentes): {stats['preserved_existing']}\n")
        f.write(f"Novas classificações: {stats['newly_classified']}\n\n")

        f.write("Distribuição por EIXO (novos):\n")
        for eixo, count in eixo_stats.most_common():
            f.write(f"  {eixo}: {count}\n")

        f.write("\nDistribuição por subcategoria (novos):\n")
        for cat, count in subcat_new.most_common():
            f.write(f"  {cat}: {count}\n")

        f.write(f"\nAtualizações:\n")
        f.write(f"  subcategoria_classificada: {write_count}\n")
        f.write(f"  recomendacao (rodada 15): {updated_rec}\n")
        f.write(f"  Excel categorias: {excel_updated_cat}\n")
        f.write(f"  Excel descrições: {excel_updated_desc}\n")

    print(f"✓ Summary exported to: {summary_path}")

    print("\n" + "=" * 70)
    print("DONE!")
    print("=" * 70)


if __name__ == '__main__':
    main()
